#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 轉 JSON 工具
使用圖形化介面選擇 Excel 檔案並轉換成 JSON 格式
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import json
import os


def select_and_convert_excel():
    """選擇 Excel 檔案並轉換成 JSON"""
    
    # 建立隱藏的 Tkinter 視窗
    root = tk.Tk()
    root.withdraw()  # 隱藏主視窗
    
    # 開啟檔案選擇對話框
    file_path = filedialog.askopenfilename(
        title="選擇要轉換的 Excel 檔案",
        filetypes=[
            ("Excel 檔案", "*.xlsx *.xls"),
            ("所有檔案", "*.*")
        ]
    )
    
    # 如果沒有選擇檔案，直接返回
    if not file_path:
        print("未選擇檔案")
        return
    
    try:
        print(f"正在處理檔案: {file_path}")
        
        # 讀取 Excel 檔案
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 取得標題列（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # 將資料轉換成 list of dict
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = value
            data.append(row_dict)
        
        # 產生 JSON 檔案名稱（與原檔案同名，但副檔名改為 .json）
        base_name = os.path.splitext(file_path)[0]
        json_path = base_name + ".json"
        
        # 寫入 JSON 檔案
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        # 顯示成功訊息
        message = f"轉換成功！\n\n共處理 {len(data)} 筆資料\n\nJSON 檔案已儲存至:\n{json_path}"
        messagebox.showinfo("轉換成功", message)
        print(message)
        
        # 詢問是否開啟檔案所在資料夾
        if messagebox.askyesno("開啟資料夾", "是否要開啟檔案所在的資料夾？"):
            # 根據作業系統開啟資料夾
            import platform
            folder_path = os.path.dirname(json_path)
            if platform.system() == "Windows":
                os.startfile(folder_path)
            elif platform.system() == "Darwin":  # macOS
                os.system(f'open "{folder_path}"')
            else:  # Linux
                os.system(f'xdg-open "{folder_path}"')
        
    except Exception as e:
        error_message = f"轉換失敗！\n\n錯誤訊息:\n{str(e)}"
        messagebox.showerror("錯誤", error_message)
        print(error_message)
    
    finally:
        root.destroy()


if __name__ == "__main__":
    print("=" * 50)
    print("Excel 轉 JSON 工具")
    print("=" * 50)
    select_and_convert_excel()
    print("\n程式執行完畢")
