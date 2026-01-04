import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime

# --- 設定 Excel 工作表名稱 ---
SHEET_NAME = 'New Words' 

# --- 基礎欄位對照 (不包含動態的 "分類" 處理) ---
COLUMN_MAPPING = {
    "Words": "Words",
    "pronunciation-1": "pronunciation-1",
    "pronunciation-2": "pronunciation-2",
    "traditional Chinese": "traditional Chinese",
    "English meaning": "English meaning",
    "等級": "等級"
    # "分類" 會在程式碼中特殊處理，支援 分類, 分類1, 分類2...
}

def select_file(title, file_type_desc="JSON Files", file_ext="*.json"):
    """通用檔案選取函式"""
    print(f"⏳ 請選擇: {title}...")
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[(file_type_desc, file_ext), ("All Files", "*.*")]
    )
    if file_path:
        print(f"📂 已選取: {file_path}")
        return file_path
    else:
        print("❌ 取消選取。")
        return None

def load_json(path):
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ JSON 讀取錯誤 {path}: {e}")
        return None

def merge_data(primary_data, secondary_data):
    """
    將次要資料 (LocalStorage) 合併入主要後台資料
    回傳: (合併後的 list, 報告內容 dict)
    """
    print("🔄 開始分析並整合資料...")
    
    # 1. 準備主要資料庫 (以 Words 為 Key)
    target_list = primary_data.get("New Words", [])
    data_map = {item.get("Words", "").strip(): item for item in target_list if "Words" in item}
    
    report = {
        "new": [],
        "updated": [],  # 格式: {"word": "abc", "diff": ["meaning changed", "category added"]}
        "total_before": len(target_list),
        "total_after": 0
    }
    
    # 2. 獲取次要資料 (customWords)
    custom_words = secondary_data.get("customWords", {})
    
    # 3. 遍歷並合併
    for word_key, word_obj in custom_words.items():
        clean_key = word_key.strip()
        
        # 移除不需要的 metadata (如 lastModified) 避免干擾比對，視情況保留
        # 這裡直接使用 word_obj 更新

        if clean_key in data_map:
            # --- 更新現有單字 ---
            original = data_map[clean_key]
            changes = []
            
            # 詳細比對每個欄位
            for k, v in word_obj.items():
                if k in ["lastModified", "UserCustom"]: continue # 忽略這些欄位的變動報告
                
                old_val = original.get(k)
                
                # 特殊處理 List 類型 (分類)
                if isinstance(v, list) and isinstance(old_val, list):
                    # 將 list 轉為 set 比對內容是否不同，或者直接比對順序
                    if v != old_val:
                        changes.append(f"[{k}] 更新")
                elif v != old_val:
                    # 一般字串比對
                     # 如果舊值是 None 或空，新值有東西 -> 新增內容
                    if not old_val and v:
                        changes.append(f"[{k}] 填入內容")
                    # 如果都有值但不一樣 -> 修改內容
                    elif old_val and v and str(old_val).strip() != str(v).strip():
                        changes.append(f"[{k}] 變更內容")
            
            if changes:
                data_map[clean_key].update(word_obj)
                report["updated"].append({"word": clean_key, "diff": ", ".join(changes)})
        else:
            # --- 新增單字 ---
            data_map[clean_key] = word_obj
            report["new"].append(clean_key)
    
    merged_list = list(data_map.values())
    report["total_after"] = len(merged_list)
    
    return merged_list, report

def copy_style(source_cell, target_cell):
    """複製 Excel 樣式"""
    if source_cell.has_style:
        try:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            pass

def ensure_headers_exist(ws, max_category_count):
    """
    確保 Excel 標題列包含 "分類", "分類1", "分類2"...
    回傳更新後的 header_map
    """
    # 讀取目前所有標題
    current_headers = {}
    last_col_idx = ws.max_column
    
    for col in range(1, last_col_idx + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            current_headers[str(val).strip()] = col
            
    # 檢查需要的分類欄位是否存在，不存在則往後新增
    # 分類 (index 0), 分類1 (index 1), 分類2 (index 2)...
    for i in range(max_category_count):
        header_name = "分類" if i == 0 else f"分類{i}"
        
        if header_name not in current_headers:
            last_col_idx += 1
            new_cell = ws.cell(row=1, column=last_col_idx)
            new_cell.value = header_name
            current_headers[header_name] = last_col_idx
            print(f"   ➕ 自動新增 Excel 欄位: {header_name}")
            
    return current_headers

def save_to_excel(data_list, output_path, template_path=None):
    """
    儲存為 Excel，支援動態分類欄位展開
    """
    print(f"📊 正在準備 Excel 檔案: {output_path}")
    
    wb = None
    ws = None
    
    # 1. 計算資料中 "分類" 的最大長度，決定需要多少個分類欄位
    max_cat_len = 0
    for item in data_list:
        cats = item.get("分類", [])
        if isinstance(cats, list):
            max_cat_len = max(max_cat_len, len(cats))
    
    # 嘗試讀取範本
    if template_path and os.path.exists(template_path):
        try:
            print(f"   ↳ 讀取範本: {template_path}")
            wb = load_workbook(template_path)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.create_sheet(SHEET_NAME)
        except Exception as e:
            print(f"   ⚠️ 範本讀取失敗，建立新檔。")
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # 2. 處理標題 (確保 分類, 分類1... 存在)
    if ws.max_row < 1 or not ws.cell(row=1, column=1).value:
        # 新檔：寫入預設標題
        headers = list(COLUMN_MAPPING.values())
        # 補上預設的分類欄位
        if max_cat_len > 0:
            headers.append("分類")
            for i in range(1, max_cat_len):
                headers.append(f"分類{i}")
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            
    # 確保標題足夠容納所有分類
    header_map = ensure_headers_exist(ws, max_cat_len)

    # 3. 建立現有資料 Row 索引
    excel_row_map = {}
    id_col_idx = header_map.get("Words")
    
    if id_col_idx:
        for row in range(2, ws.max_row + 1):
            w = ws.cell(row=row, column=id_col_idx).value
            if w:
                excel_row_map[str(w).strip()] = row

    # 4. 寫入資料
    for item in data_list:
        word_key = item.get("Words", "").strip()
        if not word_key: continue
        
        target_row = excel_row_map.get(word_key)
        is_new_row = False
        
        if not target_row:
            target_row = ws.max_row + 1
            excel_row_map[word_key] = target_row
            is_new_row = True
        
        # (A) 寫入固定欄位
        for json_key, excel_header in COLUMN_MAPPING.items():
            if excel_header in header_map:
                cell = ws.cell(row=target_row, column=header_map[excel_header])
                cell.value = item.get(json_key, "")
                
                if is_new_row and target_row > 2:
                    copy_style(ws.cell(row=target_row-1, column=header_map[excel_header]), cell)

        # (B) 特殊處理 "分類" (List -> 多個欄位)
        # JSON: ["新增", "", "UserCustom"]
        # Excel: 分類="新增", 分類1="", 分類2="UserCustom"
        categories = item.get("分類", [])
        if isinstance(categories, list):
            for i, cat_val in enumerate(categories):
                header_name = "分類" if i == 0 else f"分類{i}"
                if header_name in header_map:
                    col_idx = header_map[header_name]
                    cell = ws.cell(row=target_row, column=col_idx)
                    cell.value = cat_val
                    
                    if is_new_row and target_row > 2:
                         copy_style(ws.cell(row=target_row-1, column=col_idx), cell)

    try:
        wb.save(output_path)
        print("   ✅ Excel 儲存成功。")
    except Exception as e:
        print(f"   ❌ Excel 儲存失敗 (請先關閉檔案): {e}")

def generate_report(report_data, output_path, primary_file, secondary_file):
    """產生詳細報告"""
    print(f"📝 正在產生報告: {output_path}")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    content = []
    content.append("="*50)
    content.append(f"單字整合詳細報告")
    content.append(f"日期: {timestamp}")
    content.append("="*50)
    content.append(f"主檔 (Base): {os.path.basename(primary_file)}")
    content.append(f"更新檔 (Update): {os.path.basename(secondary_file)}")
    content.append("-" * 50)
    content.append(f"整合前總數: {report_data['total_before']}")
    content.append(f"整合後總數: {report_data['total_after']}")
    content.append(f"新增單字數: {len(report_data['new'])}")
    content.append(f"更新單字數: {len(report_data['updated'])}")
    content.append("-" * 50)
    
    # 1. 新增區塊
    content.append("\n【 🟢 新增的單字 】")
    if report_data['new']:
        for w in report_data['new']:
            content.append(f" + {w}")
    else:
        content.append(" (無)")

    # 2. 更新區塊 (包含詳細內容)
    content.append("\n【 🟡 更新的單字與細節 】")
    if report_data['updated']:
        for item in report_data['updated']:
            # item 格式: {'word': 'abc', 'diff': '...'}
            content.append(f" * {item['word']}")
            content.append(f"    └─ 變更: {item['diff']}")
    else:
        content.append(" (無)")
        
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(content))
        print("   ✅ 報告產生成功。")
    except Exception as e:
        print(f"   ❌ 報告寫入失敗: {e}")

# --- 主程式 ---
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw() 

    # 1. 選擇檔案 (順序：先選 LocalStorage 匯出檔，再選後台主檔)
    secondary_path = select_file("步驟 1/2: 請選擇【次要 JSON】(LocalStorage 匯出檔)")
    if not secondary_path:
        exit()
        
    primary_path = select_file("步驟 2/2: 請選擇【主要後台 JSON】(資料庫主檔)")
    if not primary_path:
        exit()

    # 2. 讀取與整合
    primary_data = load_json(primary_path)
    secondary_data = load_json(secondary_path)

    if primary_data and secondary_data:
        merged_list, report = merge_data(primary_data, secondary_data)
        
        # 3. 設定輸出路徑
        base_dir = os.path.dirname(primary_path)
        base_name = os.path.splitext(os.path.basename(primary_path))[0]
        
        output_json = os.path.join(base_dir, f"{base_name}_merged.json")
        output_excel = os.path.join(base_dir, f"{base_name}_merged.xlsx")
        output_report = os.path.join(base_dir, f"{base_name}_report.txt")
        
        # 尋找原始 Excel 作為範本
        potential_template = os.path.join(base_dir, f"{base_name}.xlsx")
        
        # 4. 寫入檔案
        # JSON
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump({"New Words": merged_list}, f, ensure_ascii=False, indent=4)
            
        # Excel (包含分類欄位展開邏輯)
        save_to_excel(merged_list, output_excel, template_path=potential_template)
        
        # Report (包含詳細變更欄位)
        generate_report(report, output_report, primary_path, secondary_path)

        messagebox.showinfo("整合成功", 
            f"✅ 已完成！檔案儲存於：\n{base_dir}\n\n"
            f"1. {os.path.basename(output_json)}\n"
            f"2. {os.path.basename(output_excel)}\n"
            f"3. {os.path.basename(output_report)}\n\n"
            f"新增: {len(report['new'])} 筆, 更新: {len(report['updated'])} 筆"
        )
    else:
        messagebox.showerror("錯誤", "無法讀取 JSON 檔案，請檢查格式。")