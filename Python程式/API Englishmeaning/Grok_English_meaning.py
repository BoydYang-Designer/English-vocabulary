import os
import json
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# 嘗試匯入新版 SDK
try:
    from openai import OpenAI
    NEW_OPENAI_SDK = True
except Exception:
    import openai
    NEW_OPENAI_SDK = False

# 欄位名稱
DEFAULT_REVIEW_COL = "複習"
DEFAULT_WORD_COL = "Words"
DEFAULT_TRAD_COL = "traditional Chinese"
DEFAULT_ENG_COL = "English meaning"
DEFAULT_CAT1_COL = "分類1"
DEFAULT_CAT2_COL = "分類2"
DEFAULT_CAT3_COL = "分類3"
DEFAULT_PRON1_COL = "pronunciation-1"
DEFAULT_PRON2_COL = "pronunciation-2"

# 固定的分類表
CATEGORY_GUIDE = """
主分類:藝術與美學
次分類:Design 建築 藝術 工藝 形狀 顏色 時尚

主分類:行為與心理
次分類:情緒 人物特性 動作

主分類:學術與教育
次分類:教育 文學 教育文學 語言

主分類:科學與工程
次分類:技術 科技 生物 宇宙 物理 植物 科學 形狀 動作

主分類:社會與法律
次分類:法政 歷史文化 軍事 社會 商業

主分類:醫學與健康
次分類:衛生 醫學 健康 運動 人體

主分類:商業與經濟
次分類:商業 金融 經濟

主分類:自然與環境
次分類:地理 動植物 環境 天氣

主分類:日常生活
次分類:食物 交通 居家 衣物 時間

主分類:抽象概念
次分類:概念 數學 數量
"""

# 全域 token 統計
total_prompt_tokens = 0
total_completion_tokens = 0

def get_api_key():
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        try:
            with open("api_key.txt", "r") as f:
                api_key = f.read().strip()
        except FileNotFoundError:
            api_key = None
    return api_key

def init_openai_client(api_key):
    if NEW_OPENAI_SDK:
        return OpenAI(api_key=api_key)
    else:
        openai.api_key = api_key
        return openai

def find_col_index(header, col_name):
    try:
        return [str(h).lower() if h else "" for h in header].index(col_name.lower()) + 1
    except ValueError:
        return None

def generate_chinese_meaning(client, word):
    global total_prompt_tokens, total_completion_tokens
    prompt = f"""
你是一位專業的英文-中文翻譯專家，專門為語言學習者提供簡潔的繁體中文解釋。對於單字 '{word}'，請逐步思考：1. 識別其主要詞性和常見含義。2. 只輸出繁體中文意思，使用格式如 (n.) 意思1；(v.) 意思2。3. 如果有多義，列出前2-3個最常見的。4. 無額外文字。

範例：
輸入：apple
輸出：(n.) 蘋果；(n.) 蘋果公司

輸入：run
輸出：(v.) 跑步；(n.) 奔跑；(v.) 運轉
"""
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=80,
            temperature=0
        )
        usage = response.usage
        total_prompt_tokens += usage.prompt_tokens
        total_completion_tokens += usage.completion_tokens
        return response.choices[0].message.content.strip()
    else:
        response = client.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=80,
            temperature=0
        )
        return response.choices[0].message['content'].strip()

def generate_english_details(client, word):
    global total_prompt_tokens, total_completion_tokens
    prompt = f"""
你是一位經驗豐富的英語老師，專門為中學生到進階學習者解釋單字。對於單字 '{word}'，請逐步思考：1. 評估其難度（A1-B2為簡單，C1-C2為進階）。2. 識別真實常見的詞性和含義（最多3-4個）。3. 生成簡單、學習友好的解釋。4. 嚴格遵循以下格式輸出，無多餘文字。

規則：
- 如果單字簡單（A1-B2）：使用短句（<15字）、簡單詞彙、無複雜文法。
- 如果進階（C1-C2）：允許更多細節、背景或比較。
- 詞性格式：動詞顯示 base / past / past participle / present participle / 3rd person singular；形容詞顯示 positive / comparative / superlative。
- 每個含義提供2-3個例子句子，每個句子應包含6-10個單字以增加長度。
- 結尾總結：Summary: [簡短英文摘要]；Synonyms: [5-8個簡單同義詞]；Antonyms: [5-8個簡單反義詞]。

範例（簡單單字）：
Apple is a noun: it refers to a round fruit that is red or green.
E.g. I eat a fresh apple every single day.
E.g. The ripe apple fell from the tall tree.

Summary: Apple means a common fruit.

Synonyms: fruit, pear, orange, banana, berry

Antonyms: vegetable, meat, dairy

範例（進階單字）：
Ephemeral is an adjective: positive / more ephemeral / most ephemeral: it refers to something that lasts for a very short time, often used in literature.
E.g. The beauty of cherry blossoms is truly ephemeral in spring.
E.g. Fame can be quite ephemeral in the fast digital age today.

Summary: Ephemeral describes things that are temporary.

Synonyms: fleeting, transient, short-lived, momentary, evanescent, brief, passing, temporary

Antonyms: permanent, enduring, lasting, eternal, perpetual, long-term, stable, everlasting
"""
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.7
        )
        usage = response.usage
        total_prompt_tokens += usage.prompt_tokens
        total_completion_tokens += usage.completion_tokens
        return response.choices[0].message.content.strip()
    else:
        response = client.ChatCompletion.create(
            model="gpt-4-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.7
        )
        return response.choices[0].message['content'].strip()

def generate_category(client, word, meaning):
    prompt = f"""
你是一位分類專家，專門根據給定指南將英文單字分類。對於單字 '{word}'（意思："{meaning}"），請逐步思考：1. 分析其主要含義和上下文。2. 從指南中選擇最匹配的主分類和次分類（僅限指南內的）。3. 如果多義，選擇最常見的。4. 只輸出JSON格式：{{"category1": "主分類", "category2": "次分類"}}。

{CATEGORY_GUIDE}

範例：
單字：happy，意思：feeling pleasure
輸出：{{"category1": "行為與心理", "category2": "情緒"}}

單字：computer，意思：electronic device for processing data
輸出：{{"category1": "科學與工程", "category2": "技術"}}
"""
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0
        )
        result_text = response.choices[0].message.content
    else:
        response = client.ChatCompletion.create(
            model="gpt-4-turbo",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        result_text = response.choices[0].message['content']
    return json.loads(result_text)

def tts_to_mp3_new_sdk(client, text, mp3_path):
    # 把 E.g. 改成停頓符號
    clean_text = text.replace("E.g.", "[pause]")
    response = client.audio.speech.create(
        model="tts-1",
        voice="alloy",
        input=clean_text
    )
    response.stream_to_file(mp3_path)

def main():
    global total_prompt_tokens, total_completion_tokens

    api_key = get_api_key()
    if not api_key:
        print("錯誤：未找到 OpenAI API 金鑰。請設定環境變數 OPENAI_API_KEY 或建立 api_key.txt 檔案。")
        return

    client = init_openai_client(api_key)

    root = tk.Tk()
    root.withdraw()
    excel_path = filedialog.askopenfilename(title="請選擇要處理的 Excel 檔案", filetypes=[("Excel files", "*.xlsx")])

    if not excel_path:
        print("未選擇檔案，程式結束。")
        return

    input_path = Path(excel_path)
    wb = load_workbook(input_path)
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]
    print("偵測到的欄位：", header)

    review_col = find_col_index(header, DEFAULT_REVIEW_COL)
    word_col = find_col_index(header, DEFAULT_WORD_COL)
    trad_col = find_col_index(header, DEFAULT_TRAD_COL)
    eng_col = find_col_index(header, DEFAULT_ENG_COL)
    cat1_col = find_col_index(header, DEFAULT_CAT1_COL)
    cat2_col = find_col_index(header, DEFAULT_CAT2_COL)
    cat3_col = find_col_index(header, DEFAULT_CAT3_COL)
    pron1_col = find_col_index(header, DEFAULT_PRON1_COL)
    pron2_col = find_col_index(header, DEFAULT_PRON2_COL)

    if not all([review_col, word_col, trad_col, eng_col, cat1_col, cat2_col, cat3_col, pron1_col, pron2_col]):
        print("錯誤：Excel 缺少必要的欄位。請檢查欄位名稱是否正確。")
        return

    generated_mp3_files = []
    updated_count = 0

    for row in tqdm(range(2, sheet.max_row + 1), desc="處理進度"):
        review_cell = sheet.cell(row=row, column=review_col)

        if review_cell.value == "成功":
            word_cell = sheet.cell(row=row, column=word_col)
            trad_cell = sheet.cell(row=row, column=trad_col)
            eng_cell = sheet.cell(row=row, column=eng_col)
            cat1_cell = sheet.cell(row=row, column=cat1_col)
            cat2_cell = sheet.cell(row=row, column=cat2_col)
            cat3_cell = sheet.cell(row=row, column=cat3_col)
            pron1_cell = sheet.cell(row=row, column=pron1_col)
            pron2_cell = sheet.cell(row=row, column=pron2_col)

            word = word_cell.value
            if not word:
                continue

            updated_flag = False
            new_eng_generated = False

            try:
                if not trad_cell.value:
                    trad_cell.value = generate_chinese_meaning(client if NEW_OPENAI_SDK else openai, word)
                    updated_flag = True

                if not eng_cell.value:
                    eng_cell.value = generate_english_details(client if NEW_OPENAI_SDK else openai, word)
                    updated_flag = True
                    new_eng_generated = True

                if (not cat1_cell.value) or (not cat2_cell.value):
                    try:
                        meaning_text = eng_cell.value if eng_cell.value else str(word)
                        class_data = generate_category(client if NEW_OPENAI_SDK else openai, word, meaning_text)

                        if not cat1_cell.value:
                            cat1_cell.value = class_data["category1"].strip()
                            updated_flag = True
                        if not cat2_cell.value:
                            cat2_cell.value = class_data["category2"].strip()
                            updated_flag = True
                    except Exception as e:
                        print(f"分類單字 {word} 時發生錯誤: {e}")

                if new_eng_generated:
                    safe_word = str(word).replace("/", "_").replace(" ", "_")
                    mp3_path = input_path.parent / f"{safe_word} - sentence.mp3"
                    tts_to_mp3_new_sdk(client, eng_cell.value, str(mp3_path))
                    generated_mp3_files.append(mp3_path.name)

            except Exception as e:
                print(f"處理單字 {word} 時發生錯誤: {e}")

            if updated_flag:
                updated_count += 1

    wb.save(input_path)
    print("\n處理完成！")
    print(f"總共更新了 {updated_count} 筆資料。")
    print(f"總 prompt tokens: {total_prompt_tokens}")
    print(f"總 completion tokens: {total_completion_tokens}")
    print(f"總 token: {total_prompt_tokens + total_completion_tokens}")
    if generated_mp3_files:
        print("已生成的 MP3 檔案列表:")
        for f in generated_mp3_files:
            print(f"- {f}")

if __name__ == "__main__":
    main()