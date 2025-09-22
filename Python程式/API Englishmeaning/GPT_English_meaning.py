import os
import json
from pathlib import Path
from tqdm import tqdm
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# 嘗試匯入新版 SDK
try:
    from openai import OpenAI
    NEW_OPENAI_SDK = True
except Exception:
    import openai
    NEW_OPENAI_SDK = False

# 欄位名稱
DEFAULT_REVIEW_COL = "複習"
DEFAULT_WORD_COL = "Words"
DEFAULT_TRAD_COL = "traditional Chinese"
DEFAULT_ENG_COL = "English meaning"
DEFAULT_CAT1_COL = "分類1"
DEFAULT_CAT2_COL = "分類2"
DEFAULT_CAT3_COL = "分類3"
DEFAULT_PRON1_COL = "pronunciation-1"
DEFAULT_PRON2_COL = "pronunciation-2"

# 固定的分類表
CATEGORY_GUIDE = """
主分類:藝術與美學
次分類:Design 建築 藝術 工藝 形狀 顏色 時尚

主分類:行為與心理
次分類:情緒 人物特性 動作

主分類:學術與教育
次分類:教育 文學 教育文學 語言

主分類:科學與工程
次分類:技術 科技 生物 宇宙 物理 植物 科學 形狀 動作

主分類:社會與法律
次分類:法政 歷史文化 軍事 社會 商業

主分類:醫學與健康
次分類:衛生 醫學 健康 運動 人體

主分類:商業與經濟
次分類:商業 金融 經濟

主分類:自然與環境
次分類:地理 動植物 環境 天氣

主分類:日常生活
次分類:食物 交通 居家 衣物 時間

"""

# 全域 token 統計
total_prompt_tokens = 0
total_completion_tokens = 0

def get_api_key():
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        try:
            with open("api_key.txt", "r") as f:
                api_key = f.read().strip()
        except FileNotFoundError:
            api_key = None
    return api_key

def init_openai_client(api_key):
    if NEW_OPENAI_SDK:
        return OpenAI(api_key=api_key)
    else:
        openai.api_key = api_key
        return openai

def find_col_index(header, col_name):
    try:
        return [str(h).lower() if h else "" for h in header].index(col_name.lower()) + 1
    except ValueError:
        return None

def generate_chinese_meaning(client, word):
    global total_prompt_tokens, total_completion_tokens
    prompt = f"請提供單字 '{word}' 的繁體中文意思，請只回答名詞或動詞或形容詞等...與中文意思，例如 (n.) 意思。"
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=80,
            temperature=0
        )
        usage = response.usage
        total_prompt_tokens += usage.prompt_tokens
        total_completion_tokens += usage.completion_tokens
        return response.choices[0].message.content.strip()
    else:
        response = client.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=80,
            temperature=0
        )
        return response.choices[0].message['content'].strip()

def generate_english_details(client, word):
    global total_prompt_tokens, total_completion_tokens
    prompt = f"""
Provide a clear, simple, and learner-friendly explanation for the English word '{word}'.

Rules:
- Use very simple English (for language learners).
- Only include real and common parts of speech for this word (noun, verb, adjective).
- Start each section with a header line like:
  {word.capitalize()} is a noun : singular / plural : with the following meanings:
- For verbs, show base / past / past participle / present participle / 3rd person singular.
- For adjectives, show positive / comparative / superlative.
- For each meaning:
  - Give a short and clear definition.
  - Add exactly 2 example sentences starting with "E.g."
- After meanings, always add:
  Summary:
  [short and simple summary]

  Synonyms:
  [list 4–6 simple synonyms]

  Antonyms:
  [list 4–6 simple antonyms]

Format example:

Skeleton is a noun : skeleton / skeletons : with the following meanings:
1. The hard structure of bones inside a body that supports it and protects its organs.
E.g. The skeleton of a human has 206 bones.
E.g. The scientist studied the skeleton of a dinosaur.

2. A model of bones used for teaching or display.
E.g. The school has a skeleton in the science classroom.
E.g. The museum has a life-size skeleton of a whale.

3. The basic supporting structure or outline of something.
E.g. The skeleton of the story is about friendship.
E.g. The architect showed us the skeleton of the new building.

Summary:
A "skeleton" is the framework of bones inside a body, a model of bones, or the basic structure/outline of something.

Synonyms:
framework, bones, structure, outline, model

Antonyms:
flesh, fullness, thickness, body, weight
"""
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.7
        )
        usage = response.usage
        total_prompt_tokens += usage.prompt_tokens
        total_completion_tokens += usage.completion_tokens
        return response.choices[0].message.content.strip()
    else:
        response = client.ChatCompletion.create(
            model="gpt-4-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.7
        )
        return response.choices[0].message['content'].strip()

def generate_category(client, word, meaning):
    prompt = f"""
Analyze the English word '{word}'. Its meaning is: "{meaning}".
Based on the provided category guide, choose the most suitable primary and secondary categories.

{CATEGORY_GUIDE}

Provide the output in JSON format with two keys: "category1", "category2".
"""
    if NEW_OPENAI_SDK:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0
        )
        result_text = response.choices[0].message.content
    else:
        response = client.ChatCompletion.create(
            model="gpt-4-turbo",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        result_text = response.choices[0].message['content']
    return json.loads(result_text)

def tts_to_mp3_new_sdk(client, text, mp3_path):
    # 把 E.g. 改成停頓符號
    clean_text = text.replace("E.g.", "[pause]")
    response = client.audio.speech.create(
        model="tts-1",
        voice="alloy",
        input=clean_text
    )
    response.stream_to_file(mp3_path)

def main():
    global total_prompt_tokens, total_completion_tokens

    api_key = get_api_key()
    if not api_key:
        print("錯誤：未找到 OpenAI API 金鑰。請設定環境變數 OPENAI_API_KEY 或建立 api_key.txt 檔案。")
        return

    client = init_openai_client(api_key)

    root = tk.Tk()
    root.withdraw()
    excel_path = filedialog.askopenfilename(title="請選擇要處理的 Excel 檔案", filetypes=[("Excel files", "*.xlsx")])

    if not excel_path:
        print("未選擇檔案，程式結束。")
        return

    input_path = Path(excel_path)
    wb = load_workbook(input_path)
    sheet = wb.active

    header = [cell.value for cell in sheet[1]]
    print("偵測到的欄位：", header)

    review_col = find_col_index(header, DEFAULT_REVIEW_COL)
    word_col = find_col_index(header, DEFAULT_WORD_COL)
    trad_col = find_col_index(header, DEFAULT_TRAD_COL)
    eng_col = find_col_index(header, DEFAULT_ENG_COL)
    cat1_col = find_col_index(header, DEFAULT_CAT1_COL)
    cat2_col = find_col_index(header, DEFAULT_CAT2_COL)
    cat3_col = find_col_index(header, DEFAULT_CAT3_COL)
    pron1_col = find_col_index(header, DEFAULT_PRON1_COL)
    pron2_col = find_col_index(header, DEFAULT_PRON2_COL)

    if not all([review_col, word_col, trad_col, eng_col, cat1_col, cat2_col, cat3_col, pron1_col, pron2_col]):
        print("錯誤：Excel 缺少必要的欄位。請檢查欄位名稱是否正確。")
        return

    generated_mp3_files = []
    updated_count = 0

    for row in tqdm(range(2, sheet.max_row + 1), desc="處理進度"):
        review_cell = sheet.cell(row=row, column=review_col)

        if review_cell.value == "成功":
            word_cell = sheet.cell(row=row, column=word_col)
            trad_cell = sheet.cell(row=row, column=trad_col)
            eng_cell = sheet.cell(row=row, column=eng_col)
            cat1_cell = sheet.cell(row=row, column=cat1_col)
            cat2_cell = sheet.cell(row=row, column=cat2_col)
            cat3_cell = sheet.cell(row=row, column=cat3_col)
            pron1_cell = sheet.cell(row=row, column=pron1_col)
            pron2_cell = sheet.cell(row=row, column=pron2_col)

            word = word_cell.value
            if not word:
                continue

            updated_flag = False
            new_eng_generated = False

            try:
                if not trad_cell.value:
                    trad_cell.value = generate_chinese_meaning(client if NEW_OPENAI_SDK else openai, word)
                    updated_flag = True

                if not eng_cell.value:
                    eng_cell.value = generate_english_details(client if NEW_OPENAI_SDK else openai, word)
                    updated_flag = True
                    new_eng_generated = True

                if (not cat1_cell.value) or (not cat2_cell.value):
                    try:
                        meaning_text = eng_cell.value if eng_cell.value else str(word)
                        class_data = generate_category(client if NEW_OPENAI_SDK else openai, word, meaning_text)

                        if not cat1_cell.value:
                            cat1_cell.value = class_data["category1"].strip()
                            updated_flag = True
                        if not cat2_cell.value:
                            cat2_cell.value = class_data["category2"].strip()
                            updated_flag = True
                    except Exception as e:
                        print(f"分類單字 {word} 時發生錯誤: {e}")

                if new_eng_generated:
                    safe_word = str(word).replace("/", "_").replace(" ", "_")
                    mp3_path = input_path.parent / f"{safe_word} - sentence.mp3"
                    tts_to_mp3_new_sdk(client, eng_cell.value, str(mp3_path))
                    generated_mp3_files.append(mp3_path.name)

            except Exception as e:
                print(f"處理單字 {word} 時發生錯誤: {e}")

            if updated_flag:
                updated_count += 1

    wb.save(input_path)
    print("\n處理完成！")
    print(f"總共更新了 {updated_count} 筆資料。")
    print(f"總 prompt tokens: {total_prompt_tokens}")
    print(f"總 completion tokens: {total_completion_tokens}")
    print(f"總 token: {total_prompt_tokens + total_completion_tokens}")
    if generated_mp3_files:
        print("已生成的 MP3 檔案列表:")
        for f in generated_mp3_files:
            print(f"- {f}")

if __name__ == "__main__":
    main()
