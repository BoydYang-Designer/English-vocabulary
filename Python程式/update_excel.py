import pandas as pd
from openpyxl import load_workbook

def update_excel_b(excel_a_path, excel_b_path, max_suffix=3):
    # 讀取 Excel A 和 Excel B
    df_a = pd.read_excel(excel_a_path, dtype=str).rename(columns=lambda x: x.strip())
    wb = load_workbook(excel_b_path)
    ws = wb.active  # 取得 Excel B 的主要工作表

    # 讀取 Excel B 內容（確保標題行對應欄位）
    df_b = pd.DataFrame(ws.values)
    df_b.columns = df_b.iloc[0]  # 設定標題
    df_b = df_b[1:].reset_index(drop=True)  # 移除標題行

    # 確保必要欄位
    required_columns = ['Words', '等級', '分類']
    for col in required_columns:
        if col not in df_a.columns or col not in df_b.columns:
            raise ValueError(f"Excel A 或 Excel B 缺少必要欄位: {col}")

    # **解決重複值問題**
    word_info = df_a.drop_duplicates(subset=['Words'], keep='last').set_index('Words')[['等級', '分類']].to_dict(orient='index')

    # **找出 Words 欄位在 Excel B 的正確位置**
    words_col_index = list(df_b.columns).index('Words')
    等級_col_index = list(df_b.columns).index('等級')
    分類_col_index = list(df_b.columns).index('分類')

    # **統計 Excel B 內已有的單字後綴**
    word_suffix_count = {}
    for word in df_b['Words']:
        base_word = word.rstrip('-1234567890')  # 去掉後綴，只取基礎單字
        if base_word in word_info:
            word_suffix_count[base_word] = max(word_suffix_count.get(base_word, 0), int(word.split('-')[-1]) if '-' in word else 1)

    # **步驟 1：更新 Excel B 內已存在的單字**
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        word_cell = row[words_col_index]  # `Words` 欄位
        if word_cell.value in word_info:
            等級_cell = row[等級_col_index]  # `等級` 欄位
            分類_cell = row[分類_col_index]  # `分類` 欄位

            # 只更新 `等級` 和 `分類`，不影響 `音檔`
            等級_cell.value = word_info[word_cell.value]['等級']
            分類_cell.value = word_info[word_cell.value]['分類']

    # **步驟 2：新增 `Excel A` 中 `Excel B` 沒有的單字**
    new_words = []
    for word, info in word_info.items():
        existing_count = word_suffix_count.get(word, 0)
        for i in range(existing_count + 1, max_suffix + 1):
            new_word = f"{word}-{i}"
            if new_word in df_b['Words'].values:
                continue  # 如果已經有這個單字，就跳過

            # 產生新行，確保 `Words` 欄位對應正確
            new_row = [""] * len(df_b.columns)  # 先填充所有欄位為空
            new_row[words_col_index] = new_word
            new_row[等級_col_index] = info['等級']
            new_row[分類_col_index] = info['分類']

            ws.append(new_row)

            # 更新統計，確保不會新增超過 `-3`
            word_suffix_count[word] = i
            if i == max_suffix:
                break

    # **步驟 3：儲存 Excel B（保留格式與公式）**
    wb.save(excel_b_path)
    print(f"更新完成，已保留公式並覆蓋原檔案: {excel_b_path}")

# 設定檔案路徑
excel_a_path = "C:\\conver file\\Z_total_words.xlsx"
excel_b_path = "C:\\conver file\\sentence.xlsx"

# 呼叫更新函式
update_excel_b(excel_a_path, excel_b_path)
