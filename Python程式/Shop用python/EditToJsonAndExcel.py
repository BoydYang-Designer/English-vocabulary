import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from copy import copy
from datetime import datetime

# --- 設定 Excel 工作表名稱 ---
SHEET_NAME = 'New Words' 

# --- 欄位對照設定 ---
COLUMN_MAPPING = {
    "Words": "Words",
    "pronunciation-1": "pronunciation-1",
    "pronunciation-2": "pronunciation-2",
    "traditional Chinese": "traditional Chinese",
    "English meaning": "English meaning",
    "分類": "分類",
    "等級": "等級"
}

def select_file(title, file_type_desc="JSON Files", file_ext="*.json"):
    """通用檔案選取函式"""
    print(f"⏳ 請選擇: {title}...")
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[(file_type_desc, file_ext), ("All Files", "*.*")]
    )
    if file_path:
        print(f"📂 已選取: {file_path}")
        return file_path
    else:
        print("❌ 取消選取。")
        return None

def load_json(path):
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ JSON 讀取錯誤 {path}: {e}")
        return None

def merge_data(primary_data, secondary_data):
    """
    將次要資料 (LocalStorage) 合併入主要後台資料
    回傳: (合併後的 list, 報告內容 dict)
    """
    print("🔄 開始分析並整合資料...")
    
    # 1. 準備主要資料庫 (以 Words 為 Key)
    target_list = primary_data.get("New Words", [])
    data_map = {item.get("Words", "").strip(): item for item in target_list if "Words" in item}
    
    # 2. 獲取次要資料 (customWords)
    custom_words = secondary_data.get("customWords", {})
    
    report = {
        "new": [],
        "updated": [],
        "total_before": len(target_list),
        "total_after": 0
    }
    
    # 3. 遍歷並合併
    for word_key, word_obj in custom_words.items():
        clean_key = word_key.strip()
        
        # 處理資料格式 (如分類 Array -> 保留原樣，寫入 Excel 時再處理)
        if "分類" in word_obj and isinstance(word_obj["分類"], list):
            pass 

        if clean_key in data_map:
            # --- 更新 ---
            # 檢查是否有實際變更 (簡易檢查)
            original = data_map[clean_key]
            has_changed = False
            for k, v in word_obj.items():
                if original.get(k) != v:
                    has_changed = True
                    break
            
            if has_changed:
                data_map[clean_key].update(word_obj)
                report["updated"].append(clean_key)
        else:
            # --- 新增 ---
            data_map[clean_key] = word_obj
            report["new"].append(clean_key)
    
    merged_list = list(data_map.values())
    report["total_after"] = len(merged_list)
    
    return merged_list, report

def copy_style(source_cell, target_cell):
    """複製 Excel 樣式"""
    if source_cell.has_style:
        try:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            pass

def save_to_excel(data_list, output_path, template_path=None):
    """
    儲存為 Excel。
    如果 template_path 存在，會嘗試讀取它來保留樣式；否則建立新檔。
    """
    print(f"📊 正在準備 Excel 檔案: {output_path}")
    
    wb = None
    ws = None
    header_map = {}
    excel_row_map = {} # Words -> Row Index
    
    # 嘗試讀取範本 (主要 JSON 同檔名的 .xlsx) 以保留格式
    if template_path and os.path.exists(template_path):
        try:
            print(f"   ↳ 偵測到同名 Excel 範本: {template_path}，正在讀取格式...")
            wb = load_workbook(template_path)
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.create_sheet(SHEET_NAME)
        except Exception as e:
            print(f"   ⚠️ 範本讀取失敗 ({e})，將建立新檔。")
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
    else:
        print("   ↳ 無現有 Excel 範本，建立新檔。")
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # 1. 處理標題列 (若為新檔則寫入預設標題)
    if ws.max_row < 1 or not ws.cell(row=1, column=1).value:
        # 寫入預設標題
        headers = list(COLUMN_MAPPING.values())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
    
    # 建立標題索引
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            header_map[str(val).strip()] = col

    # 建立現有資料索引 (避免重複新增)
    id_col_idx = header_map.get("Words")
    if id_col_idx:
        for row in range(2, ws.max_row + 1):
            w = ws.cell(row=row, column=id_col_idx).value
            if w:
                excel_row_map[str(w).strip()] = row
    else:
        # 若找不到 Words 欄位，這張表可能有問題，視為新表處理
        pass

    # 2. 寫入資料
    for item in data_list:
        word_key = item.get("Words", "").strip()
        if not word_key: continue
        
        target_row = excel_row_map.get(word_key)
        is_new_row = False
        
        if not target_row:
            target_row = ws.max_row + 1
            excel_row_map[word_key] = target_row
            is_new_row = True
        
        # 填入各欄位
        for json_key, excel_header in COLUMN_MAPPING.items():
            if excel_header in header_map:
                col_idx = header_map[excel_header]
                cell = ws.cell(row=target_row, column=col_idx)
                
                val = item.get(json_key, "")
                # Array 轉 String
                if isinstance(val, list):
                    val = val[0] if len(val) > 0 else ""
                
                cell.value = val
                
                # 複製樣式 (若是新增行，參考上一行)
                if is_new_row and target_row > 2:
                    source_cell = ws.cell(row=target_row - 1, column=col_idx)
                    copy_style(source_cell, cell)

    try:
        wb.save(output_path)
        print("   ✅ Excel 儲存成功。")
    except Exception as e:
        print(f"   ❌ Excel 儲存失敗 (請確認檔案未被開啟): {e}")

def generate_report(report_data, output_path, primary_file, secondary_file):
    """產生整合報告文字檔"""
    print(f"📝 正在產生報告: {output_path}")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    content = []
    content.append("="*40)
    content.append(f"單字整合報告 - {timestamp}")
    content.append("="*40)
    content.append(f"主要來源 (Base): {os.path.basename(primary_file)}")
    content.append(f"次要來源 (Update): {os.path.basename(secondary_file)}")
    content.append("-" * 40)
    content.append(f"整合前總數: {report_data['total_before']}")
    content.append(f"整合後總數: {report_data['total_after']}")
    content.append(f"新增單字數: {len(report_data['new'])}")
    content.append(f"更新單字數: {len(report_data['updated'])}")
    content.append("-" * 40)
    
    if report_data['new']:
        content.append("\n[新增的單字]:")
        for w in report_data['new']:
            content.append(f" + {w}")
    else:
        content.append("\n[無新增單字]")

    if report_data['updated']:
        content.append("\n[更新的單字]:")
        for w in report_data['updated']:
            content.append(f" * {w}")
    else:
        content.append("\n[無更新單字]")
        
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(content))
        print("   ✅ 報告產生成功。")
    except Exception as e:
        print(f"   ❌ 報告寫入失敗: {e}")

# --- 主程式 ---
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw() # 隱藏主視窗

    # 1. 選擇檔案
    secondary_path = select_file("步驟 1/2: 請選擇【次要 JSON】(LocalStorage 匯出檔)")
    if not secondary_path:
        exit()
        
    primary_path = select_file("步驟 2/2: 請選擇【主要後台 JSON】(資料庫主檔)")
    if not primary_path:
        exit()

    # 2. 讀取資料
    primary_data = load_json(primary_path)
    secondary_data = load_json(secondary_path)

    if primary_data and secondary_data:
        # 3. 執行整合
        merged_list, report = merge_data(primary_data, secondary_data)
        
        # 4. 準備輸出路徑 (存放在主要 JSON 的同目錄下)
        base_dir = os.path.dirname(primary_path)
        base_name = os.path.splitext(os.path.basename(primary_path))[0]
        
        # 輸出檔名設定
        output_json_path = os.path.join(base_dir, f"{base_name}_merged.json")
        output_excel_path = os.path.join(base_dir, f"{base_name}_merged.xlsx")
        output_report_path = os.path.join(base_dir, f"{base_name}_report.txt")
        
        # 尋找是否存在同名 Excel 作為格式範本 (例如 Z_total_words.xlsx)
        potential_template = os.path.join(base_dir, f"{base_name}.xlsx")
        
        # 5. 輸出檔案
        # (A) 寫入 JSON
        print(f"💾 儲存整合 JSON: {output_json_path}")
        new_backend_data = {"New Words": merged_list}
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(new_backend_data, f, ensure_ascii=False, indent=4)
            
        # (B) 寫入 Excel
        save_to_excel(merged_list, output_excel_path, template_path=potential_template)
        
        # (C) 寫入報告
        generate_report(report, output_report_path, primary_path, secondary_path)

        # 完成提示
        messagebox.showinfo("整合完成", 
            f"成功產出以下檔案於 {base_dir}：\n\n"
            f"1. JSON: {os.path.basename(output_json_path)}\n"
            f"2. Excel: {os.path.basename(output_excel_path)}\n"
            f"3. Report: {os.path.basename(output_report_path)}\n\n"
            f"新增: {len(report['new'])} 筆, 更新: {len(report['updated'])} 筆"
        )
    else:
        messagebox.showerror("錯誤", "無法讀取 JSON 檔案，請檢查檔案格式。")