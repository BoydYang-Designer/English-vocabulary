import pandas as pd
import openpyxl
import re
import json

def extract_eg_sentences(text):
    """從 'English meaning' 欄位中提取以 'E.g.' 開頭的句子"""
    if pd.isna(text):
        return []
    sentences = re.findall(r'E\.g\.\s*(.*?)\.', text)
    return [s.strip() for s in sentences]

def get_max_suffix(word, df_b):
    """獲取 Excel B 中該單字的最大編號"""
    pattern = rf'^{word}-(\d+)$'
    suffixes = df_b['Words'].str.extract(pattern).dropna().astype(int)
    return suffixes.max().item() if not suffixes.empty else 0

def preprocess_sentence(sentence):
    """預處理句子：去除標點符號並轉換為小寫"""
    sentence = re.sub(r'[^\w\s]', '', sentence)  # 去除標點符號
    return sentence.lower().strip()  # 轉換為小寫並去除多餘空白

def check_duplicate_words(excel_a_path, excel_b_path, updated_excel_b_path):
    """檢查 Excel A、Excel B 和更新後的 Excel B 中 Words 欄位的重複單字"""
    # 讀取三個 Excel 文件
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    df_updated_b = pd.read_excel(updated_excel_b_path)
    
    # 獲取 Words 欄位的重複情況
    duplicates_a = df_a['Words'].value_counts()
    duplicates_b = df_b['Words'].value_counts()
    duplicates_updated_b = df_updated_b['Words'].value_counts()
    
    # 過濾出重複的單字（出現次數大於 1）
    dup_a = duplicates_a[duplicates_a > 1].to_dict()
    dup_b = duplicates_b[duplicates_b > 1].to_dict()
    dup_updated_b = duplicates_updated_b[duplicates_updated_b > 1].to_dict()
    
    # 構建結果
    duplicate_summary = {
        'Excel_A': [{'Word': word, '出現次數': count} for word, count in dup_a.items()],
        'Excel_B': [{'Word': word, '出現次數': count} for word, count in dup_b.items()],
        'Updated_Excel_B': [{'Word': word, '出現次數': count} for word, count in dup_updated_b.items()]
    }
    
    return duplicate_summary

def update_excel_b(excel_a_path, excel_b_path, output_path, json_path):
    """將 Excel A 的新例句更新到 Excel B，並更新現有記錄的分類和等級，並生成 JSON 記錄"""
    # 讀取 Excel A 和 Excel B
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    # 載入 Excel B 的工作簿以保留格式
    wb = openpyxl.load_workbook(excel_b_path)
    ws = wb.active
    
    # 儲存新條目和更新記錄
    new_rows = []
    updated_records = []
    updated_cells = 0  # 記錄更新的單元格數量
    
    for index, row in df_a.iterrows():
        word = row['Words']
        eg_sentences = extract_eg_sentences(row['English meaning'])
        category_a = row['分類'] if pd.notna(row['分類']) else ''
        level_a = row['等級'] if pd.notna(row['等級']) else ''
        
        # 獲取 Excel B 中該單字的記錄
        word_records = df_b[df_b['Words'].str.startswith(word + '-')]
        max_suffix = get_max_suffix(word, df_b)
        
        # 預處理現有句子並存為集合
        existing_sentences = set(preprocess_sentence(s) for s in word_records['句子'].dropna())
        
        # 更新現有記錄的分類和等級
        for b_index, b_row in word_records.iterrows():
            b_row_idx = b_index + 2  # Excel 行號從 1 開始，且有標題行
            update_needed = False
            updates = {}
            
            # 檢查分類
            current_category = b_row['分類'] if pd.notna(b_row['分類']) else ''
            if not current_category and category_a:
                ws.cell(row=b_row_idx, column=3, value=category_a)  # 分類在第3列
                updates['分類'] = category_a
                update_needed = True
            
            # 檢查等級
            current_level = b_row['等級'] if pd.notna(b_row['等級']) else ''
            if not current_level and level_a:
                ws.cell(row=b_row_idx, column=2, value=level_a)  # 等級在第2列
                updates['等級'] = level_a
                update_needed = True
            
            if update_needed:
                updated_records.append({
                    'Words': b_row['Words'],
                    '更新': updates
                })
                updated_cells += 1
        
        # 處理新例句
        if eg_sentences:
            for sentence in eg_sentences:
                preprocessed_sentence = preprocess_sentence(sentence)
                if preprocessed_sentence not in existing_sentences:
                    max_suffix += 1
                    new_word = f"{word}-{max_suffix}"
                    new_row = {
                        '音檔': '',
                        '等級': level_a,
                        '分類': category_a,
                        'Words': new_word,
                        '名人': '',
                        '句子': sentence,  # 保留原始句子
                        '中文': ''
                    }
                    new_rows.append(new_row)
                    existing_sentences.add(preprocessed_sentence)
    
    # 追加新條目到 Excel
    if new_rows:
        for row in pd.DataFrame(new_rows).itertuples(index=False):
            ws.append(row)
    
    # 保存更新後的 Excel 文件
    wb.save(output_path)
    
    # 準備 JSON 數據
    update_summary = {
        '新記錄': [
            {'Words': row['Words'], '句子': row['句子']}
            for row in new_rows
        ],
        '更新記錄': updated_records
    }
    
    # 檢查重複單字
    duplicate_summary = check_duplicate_words(excel_a_path, excel_b_path, output_path)
    update_summary['重複單字'] = duplicate_summary
    
    # 將更新內容寫入 JSON 檔
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(update_summary, json_file, ensure_ascii=False, indent=4)
    
    # 輸出訊息
    if new_rows or updated_cells:
        print(f"已更新 {len(new_rows)} 條新記錄和 {updated_cells} 個現有記錄的分類/等級到 {output_path}")
    else:
        print("沒有新記錄或更新需要保存")
    print(f"更新詳細記錄及重複單字分析已保存到 {json_path}")

# 使用示例
excel_a_path = 'Z_total_words.xlsx'
excel_b_path = 'sentence.xlsx'
output_path = 'updated_sentence.xlsx'
json_path = 'update_summary.json'
update_excel_b(excel_a_path, excel_b_path, output_path, json_path)