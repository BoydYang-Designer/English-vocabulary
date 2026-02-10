import pandas as pd
import openpyxl
import re
import json
import os
import sys
import tkinter as tk
import traceback  # 新增：用於顯示詳細錯誤訊息
from tkinter import filedialog, messagebox

def extract_eg_sentences(text):
    """從 'English meaning' 欄位中提取以 'E.g.' 開頭的句子，處理空格數量誤植"""
    if pd.isna(text):
        return []
    sentences = re.findall(r'E\.g\.\s*(.*?)(?:\.\s*$|\.\s+|\n|$)', text, re.MULTILINE)
    cleaned_sentences = [s.strip() for s in sentences if s.strip()]
    # print(f"提取的例句: {cleaned_sentences}") # 除錯用，可註解掉減少干擾
    return cleaned_sentences

def get_max_suffix(word, df_b):
    """獲取 Excel B 中該單字的最大編號"""
    pattern = rf'^{re.escape(word)}-(\d+)$'
    suffixes = df_b['Words'].str.extract(pattern).dropna()
    if suffixes.empty:
        return 0
    try:
        suffixes = suffixes.astype(int)
        return int(suffixes.max().item())
    except Exception as e:
        print(f"處理單字 {word} 的編號時發生錯誤: {e}")
        print(f"問題數據: {suffixes}")
        raise

def preprocess_sentence(sentence):
    """預處理句子：去除標點符號並轉換為小寫"""
    sentence = re.sub(r'[^\w\s]', '', sentence)
    return sentence.lower().strip()

def update_excel_b(excel_a_path, excel_b_path, output_path):
    """將 Excel A 的新例句更新到 Excel B，並更新現有記錄的分類和等級"""
    print(f"正在讀取檔案 A: {excel_a_path}")
    if not os.path.exists(excel_a_path):
        raise FileNotFoundError(f"找不到 Excel A 檔案: {excel_a_path}")
    
    print(f"正在讀取檔案 B: {excel_b_path}")
    if not os.path.exists(excel_b_path):
        raise FileNotFoundError(f"找不到 Excel B 檔案: {excel_b_path}")
    
    # 獲取原始檔案大小
    input_size = os.path.getsize(excel_b_path) / 1024  # 單位：KB
    print(f"Excel B 原始檔案大小: {input_size:.2f} KB")
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    print(f"Excel A 行數: {len(df_a)}")
    print(f"Excel B 原始行數: {len(df_b)}")
    
    wb = openpyxl.load_workbook(excel_b_path)
    ws = wb.active
    
    new_rows = []
    
    print("開始比對並更新資料...")
    for index, row in df_a.iterrows():
        word = row['Words']
        if pd.isna(word) or not isinstance(word, str):
            continue
        
        eg_sentences = extract_eg_sentences(row['English meaning'])

        # 🔹 支援多分類
        categories = []
        for c in ['分類1', '分類2', '分類3']:
            if c in row and pd.notna(row[c]):
                categories.append(str(row[c]).strip())
        category1 = categories[0] if len(categories) > 0 else ''
        category2 = categories[1] if len(categories) > 1 else ''
        category3 = categories[2] if len(categories) > 2 else ''

        level_a = row['等級'] if pd.notna(row['等級']) else ''
        
        word_records = df_b[df_b['Words'].notna() & df_b['Words'].str.match(rf'^{re.escape(word)}-\d+$')]
        max_suffix = get_max_suffix(word, df_b)
        
        existing_sentences = set(preprocess_sentence(s) for s in word_records['句子'].dropna())
        
        # 更新現有資料
        for b_index, b_row in word_records.iterrows():
            b_row_idx = b_index + 2
            update_needed = False
            
            # 更新分類
            if category1 or category2 or category3:
                ws.cell(row=b_row_idx, column=3, value=category1)
                ws.cell(row=b_row_idx, column=4, value=category2)
                ws.cell(row=b_row_idx, column=5, value=category3)
                update_needed = True
            
            current_level = b_row['等級'] if pd.notna(b_row['等級']) else ''
            if not current_level and level_a:
                ws.cell(row=b_row_idx, column=2, value=level_a)
                update_needed = True
                
            if update_needed:
                print(f"更新: {b_row['Words']} 分類/等級已同步")
        
        # 新增例句
        if eg_sentences:
            for sentence in eg_sentences:
                preprocessed_sentence = preprocess_sentence(sentence)
                if preprocessed_sentence not in existing_sentences:
                    max_suffix += 1
                    new_word = f"{word}-{max_suffix}"
                    new_row = {
                        '音檔': '',
                        '等級': level_a,
                        '分類1': category1,
                        '分類2': category2,
                        '分類3': category3,
                        'Words': new_word,
                        '名人': '',
                        '句子': sentence,
                        '中文': ''
                    }
                    new_rows.append(new_row)
                    existing_sentences.add(preprocessed_sentence)
                    print(f"++ 添加新例句: {new_word}")
    
    if new_rows:
        for row in pd.DataFrame(new_rows).itertuples(index=False):
            ws.append(row)
    
    print(f"本次總共新增行數: {len(new_rows)}")
    
    wb.save(output_path)
    
    df_output = pd.read_excel(output_path)
    output_size = os.path.getsize(output_path) / 1024  # 單位：KB
    print(f"Excel B 輸出行數: {len(df_output)}")
    print(f"Excel B 輸出檔案大小: {output_size:.2f} KB")
    
    if len(df_output) < len(df_b):
        print("警告：輸出行數少於原始行數，可能有數據丟失！")
    
    print(f"已更新 Excel B，結果保存到 {output_path}")
    return True

def compare_excel_files(excel_a_path, excel_b_path, output_json_path):
    """比對兩個 Excel 檔案的指定欄位（不含音檔），並在有差異時記錄到 JSON 檔案"""
    print("\n正在執行差異比對...")
    columns_to_compare = ['等級', '分類1', '分類2', '分類3', 'Words', '名人', '句子', '中文']
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    for col in columns_to_compare:
        if col not in df_a.columns:
            df_a[col] = ''
        if col not in df_b.columns:
            df_b[col] = ''
    
    df_a = df_a.fillna('')
    df_b = df_b.fillna('')
    
    a_dict = {row['Words']: row.to_dict() for _, row in df_a.iterrows()}
    b_dict = {row['Words']: row.to_dict() for _, row in df_b.iterrows()}
    
    differences = []
    
    common_words = set(a_dict.keys()) & set(b_dict.keys())
    for word in common_words:
        a_row = a_dict[word]
        b_row = b_dict[word]
        diff = {}
        
        for col in columns_to_compare:
            if col != 'Words':
                a_val = a_row[col]
                b_val = b_row[col]
                if a_val != b_val:
                    diff[col] = {'A 值': a_val, 'B 值': b_val}
        
        if diff:
            differences.append({
                'Words': word,
                '狀態': 'A 和 B 均有，但內容有差異',
                '差異內容': diff
            })
    
    a_only_words = set(a_dict.keys()) - set(b_dict.keys())
    for word in a_only_words:
        differences.append({
            'Words': word,
            '狀態': 'A 有，B 無',
            'A 內容': {k: v for k, v in a_dict[word].items() if k in columns_to_compare}
        })
    
    b_only_words = set(b_dict.keys()) - set(a_dict.keys())
    for word in b_only_words:
        differences.append({
            'Words': word,
            '狀態': 'B 有，A 無',
            'B 內容': {k: v for k, v in b_dict[word].items() if k in columns_to_compare}
        })
    
    if differences:
        with open(output_json_path, 'w', encoding='utf-8') as json_file:
            json.dump(differences, json_file, ensure_ascii=False, indent=4)
        print(f"比對完成，差異已記錄到 {output_json_path}")
    else:
        print("A 和 B 之間沒有差異，未生成 JSON 檔案")

def select_files():
    """使用圖形介面選擇檔案"""
    print("\n" + "="*60)
    print("程式啟動中... 請稍候")
    print("="*60)
    
    root = tk.Tk()
    root.withdraw()  # 隱藏主視窗
    
    # 確保視窗顯示在最上層
    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    
    print("\n等待使用者選擇檔案 (請留意跳出的視窗)...")
    
    # 顯示歡迎訊息
    messagebox.showinfo(
        "Excel 更新程式", 
        "歡迎使用 Excel 更新程式！\n\n" +
        "接下來將會依序要求您選擇：\n" +
        "1. Excel A 檔案 (來源檔案)\n" +
        "2. Excel B 檔案 (要更新的檔案)\n\n" +
        "按確定開始..."
    )
    
    print("顯示 Excel A 選擇對話框...")
    
    # 選擇 Excel A 檔案
    excel_a_path = filedialog.askopenfilename(
        title="步驟 1/2：選擇 Excel A 檔案 (來源)",
        filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")],
        parent=root
    )
    
    if not excel_a_path:
        print("使用者取消選擇 Excel A")
        root.destroy()
        return None, None
    
    print(f"已選擇 Excel A: {excel_a_path}")
    
    print("顯示 Excel B 選擇對話框...")
    
    # 選擇 Excel B 檔案
    excel_b_path = filedialog.askopenfilename(
        title="步驟 2/2：選擇 Excel B 檔案 (目標)",
        filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")],
        parent=root
    )
    
    if not excel_b_path:
        print("使用者取消選擇 Excel B")
        root.destroy()
        return None, None
    
    print(f"已選擇 Excel B: {excel_b_path}")
    
    root.destroy()
    
    return excel_a_path, excel_b_path

def main():
    """主函數邏輯"""
    
    # 使用圖形介面選擇檔案
    excel_a_path, excel_b_path = select_files()
    
    if not excel_a_path or not excel_b_path:
        print("\n未選擇檔案，操作已取消。")
        return # 結束 main，但不關閉視窗，因為外面有 finally
    
    # 取得 Excel B 所在的目錄，輸出檔案將保存在同一目錄
    output_dir = os.path.dirname(excel_b_path)
    output_path = os.path.join(output_dir, 'updated_sentence.xlsx')
    output_json_path = os.path.join(output_dir, 'comparison_result.json')
    
    print(f"\n輸出目錄: {output_dir}")
    print(f"預計輸出檔案: updated_sentence.xlsx")
    
    # 1. 驗證 Excel B 格式
    print("\n開始驗證 Excel B 格式...")
    try:
        df_b = pd.read_excel(excel_b_path)
        # 簡單檢查必要欄位
        if 'Words' not in df_b.columns:
             raise ValueError("Excel B 檔案缺少必要的 'Words' 欄位！")

        invalid_words = df_b[df_b['Words'].notna() & ~df_b['Words'].str.match(r'^[\w\s\''éèêëáàâãäåíìîïóòôõöúùûüýÿ-]+-\d+$', na=False)]
        
        if not invalid_words.empty:
            print("警告：發現部分 Words 欄位值格式可能不符，但程式將嘗試繼續執行。")
            # 不強制 return，只做警告
        
        print("✓ Excel B 格式驗證通過")
        
    except Exception as e:
        print(f"讀取 Excel B 時發生嚴重錯誤: {e}")
        traceback.print_exc()
        return

    # 2. 執行更新
    print("\n開始更新 Excel B...")
    update_excel_b(excel_a_path, excel_b_path, output_path)
    
    # 3. 執行比對
    print("\n開始比對檔案...")
    compare_excel_files(excel_b_path, output_path, output_json_path)
    
    success_msg = (
        f"✓ 處理完成！\n\n"
        f"輸出檔案位置：\n"
        f"• {output_path}\n"
        f"• {output_json_path}\n\n"
        f"請到該目錄查看結果。"
    )
    
    print("\n" + "="*60)
    print("全部處理完成！")
    print("="*60)
    
    # 最後彈窗通知
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    messagebox.showinfo("完成", success_msg)
    root.destroy()

# ==========================================
# 程式入口點：防止閃退的關鍵設計
# ==========================================
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 如果程式崩潰，會捕捉到這裡
        print("\n" + "!"*60)
        print("程式發生錯誤 (Error):")
        print(f"{e}")
        print("\n詳細錯誤訊息 (Traceback):")
        traceback.print_exc()
        print("!"*60)
    finally:
        # 無論成功或失敗，這行最後一定會執行
        print("\n" + "-"*30)
        input("請按 Enter 鍵關閉視窗...")