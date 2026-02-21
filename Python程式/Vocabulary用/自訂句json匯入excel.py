#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
custom_sentences_merge.py
自訂句子 JSON 整合到 Excel 工具

功能：
  1. 選擇自訂句子 JSON（從網頁匯出）
  2. 選擇原始 Excel 檔案
  3. 比對重複（Words + 句子 聯合 key）
  4. 合併後按 Words 字母排序
  5. 覆蓋原始 Excel 檔案
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ===== 主要工作函式 =====

def normalize(text):
    """標準化字串以利比對"""
    if text is None:
        return ""
    return str(text).strip().lower()


def merge_json_into_excel(json_path, excel_path, progress_callback=None):
    """
    合併 JSON 自訂句子到 Excel
    回傳: (added_count, skipped_count, error_message)
    """
    try:
        # 讀取 JSON
        with open(json_path, 'r', encoding='utf-8') as f:
            custom_data = json.load(f)
        
        if not isinstance(custom_data, list):
            return 0, 0, "JSON 格式錯誤：應為陣列（list）格式"
        
        if progress_callback:
            progress_callback(f"✅ 讀取 JSON：{len(custom_data)} 筆資料")

        # 讀取 Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        # 取得標題列
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['音檔', '等級', '分類1', '分類2', '分類3', 'Words', '名人', '句子', '中文', '記錄']
        
        # 確認欄位存在，若不存在則補齊
        for h in expected_headers:
            if h not in headers:
                headers.append(h)

        if progress_callback:
            progress_callback(f"✅ 讀取 Excel：{ws.max_row - 1} 筆現有資料")

        # 建立現有資料的 key set（Words + 句子）
        existing_keys = set()
        existing_words_sentences = {}  # Words -> [sentence1, sentence2, ...]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val
            
            words_val = normalize(row_dict.get('Words', ''))
            sentence_val = normalize(row_dict.get('句子', ''))
            
            if words_val:
                exact_key = f"{words_val}|||{sentence_val}"
                existing_keys.add(exact_key)
                
                if words_val not in existing_words_sentences:
                    existing_words_sentences[words_val] = []
                existing_words_sentences[words_val].append(sentence_val)

        # 比對並收集要新增的資料
        rows_to_add = []
        skipped = []
        
        for item in custom_data:
            words_val = normalize(item.get('Words', ''))
            sentence_val = normalize(item.get('句子', ''))
            exact_key = f"{words_val}|||{sentence_val}"
            
            if exact_key in existing_keys:
                skipped.append(item.get('Words', ''))
                if progress_callback:
                    progress_callback(f"⏭ 跳過重複：{item.get('Words', '')} | {str(item.get('句子', ''))[:40]}...")
            else:
                rows_to_add.append(item)

        if progress_callback:
            progress_callback(f"\n📊 分析完成：{len(rows_to_add)} 筆新增 / {len(skipped)} 筆跳過")

        if len(rows_to_add) == 0:
            if progress_callback:
                progress_callback("\n✅ 沒有需要新增的資料（全部已存在）")
            wb.close()
            return 0, len(skipped), None

        # 新增資料到工作表
        for item in rows_to_add:
            new_row = []
            for h in expected_headers:
                val = item.get(h, '')
                # 音檔欄位保持空白
                if h == '音檔':
                    val = ''
                # 移除 isCustom、createdAt 等非 Excel 欄位
                new_row.append(val if val != '' else None)
            ws.append(new_row)
            if progress_callback:
                progress_callback(f"➕ 新增：{item.get('Words', '')}")

        if progress_callback:
            progress_callback("\n🔤 開始按 Words 排序...")

        # 重新排序（跳過標題列）
        all_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_rows.append(list(row))

        def sort_key(row):
            """取得 Words 欄位的值作為排序 key"""
            words_idx = headers.index('Words') if 'Words' in headers else 5
            val = row[words_idx] if words_idx < len(row) else ''
            return str(val).lower() if val else 'zzz'

        all_rows.sort(key=sort_key)

        # 清除舊資料（保留標題）
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # 縮減工作表至資料長度（避免殘留空行）
        if ws.max_row > len(all_rows) + 1:
            ws.delete_rows(len(all_rows) + 2, ws.max_row - len(all_rows) - 1)

        # 寫回排序後的資料
        for i, row_data in enumerate(all_rows, start=2):
            for j, val in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=val)

        # 覆蓋儲存
        wb.save(excel_path)
        wb.close()

        if progress_callback:
            progress_callback(f"\n✅ 儲存完成！檔案：{excel_path}")
            progress_callback(f"📊 最終統計：新增 {len(rows_to_add)} 筆 / 跳過 {len(skipped)} 筆")

        return len(rows_to_add), len(skipped), None

    except Exception as e:
        return 0, 0, str(e)


# ===== GUI 主視窗 =====

class MergeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("自訂句子整合工具")
        self.root.geometry("560x640")
        self.root.resizable(True, True)
        self.root.configure(bg="#FDFBF6")

        self.json_path = tk.StringVar()
        self.excel_path = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        # 標題
        title_frame = tk.Frame(self.root, bg="#A1887F", pady=12)
        title_frame.pack(fill=tk.X)
        tk.Label(
            title_frame, text="📚 自訂句子整合工具",
            font=("Arial", 16, "bold"), bg="#A1887F", fg="white"
        ).pack()
        tk.Label(
            title_frame, text="將自訂 JSON 合併至原始 Excel",
            font=("Arial", 11), bg="#A1887F", fg="#F5E6D3"
        ).pack()

        main_frame = tk.Frame(self.root, bg="#FDFBF6", padx=20, pady=15)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ---- 步驟1：選擇 JSON ----
        self._section_label(main_frame, "步驟 1：選擇自訂句子 JSON 檔案")
        json_frame = tk.Frame(main_frame, bg="#FDFBF6")
        json_frame.pack(fill=tk.X, pady=(4, 10))
        tk.Entry(
            json_frame, textvariable=self.json_path,
            font=("Arial", 11), bg="white", relief=tk.SOLID, bd=1
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)
        tk.Button(
            json_frame, text="選擇", command=self._select_json,
            bg="#7986CB", fg="white", font=("Arial", 11), relief=tk.FLAT,
            padx=10, cursor="hand2"
        ).pack(side=tk.LEFT, padx=(6, 0))

        # ---- 步驟2：選擇 Excel ----
        self._section_label(main_frame, "步驟 2：選擇原始 Excel 檔案")
        excel_frame = tk.Frame(main_frame, bg="#FDFBF6")
        excel_frame.pack(fill=tk.X, pady=(4, 10))
        tk.Entry(
            excel_frame, textvariable=self.excel_path,
            font=("Arial", 11), bg="white", relief=tk.SOLID, bd=1
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)
        tk.Button(
            excel_frame, text="選擇", command=self._select_excel,
            bg="#7986CB", fg="white", font=("Arial", 11), relief=tk.FLAT,
            padx=10, cursor="hand2"
        ).pack(side=tk.LEFT, padx=(6, 0))

        # ---- 說明 ----
        info_frame = tk.Frame(main_frame, bg="#FFF8E7", relief=tk.SOLID, bd=1, pady=8, padx=12)
        info_frame.pack(fill=tk.X, pady=(0, 12))
        tk.Label(
            info_frame,
            text="ℹ️  比對規則：Words + 句子 完全相同視為重複（跳過）\n"
                 "📌  合併完成後將依 Words 字母排序並覆蓋原始 Excel",
            font=("Arial", 10), bg="#FFF8E7", fg="#795548",
            justify=tk.LEFT
        ).pack(anchor=tk.W)

        # ---- 開始按鈕 ----
        tk.Button(
            main_frame, text="▶  開始整合",
            command=self._start_merge,
            bg="#A1887F", fg="white",
            font=("Arial", 14, "bold"),
            relief=tk.FLAT, pady=10, cursor="hand2",
            activebackground="#8D6E63", activeforeground="white"
        ).pack(fill=tk.X, pady=(0, 12))

        # ---- 進度日誌 ----
        self._section_label(main_frame, "執行日誌")
        log_frame = tk.Frame(main_frame, bg="#FDFBF6")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        
        scrollbar = tk.Scrollbar(log_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(
            log_frame, height=12,
            font=("Courier", 10), bg="#1e1e1e", fg="#d4d4d4",
            relief=tk.SOLID, bd=1,
            yscrollcommand=scrollbar.set,
            state=tk.DISABLED
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)

        # ---- 底部按鈕 ----
        bottom_frame = tk.Frame(main_frame, bg="#FDFBF6")
        bottom_frame.pack(fill=tk.X, pady=(8, 0))
        tk.Button(
            bottom_frame, text="🗑 清除日誌", command=self._clear_log,
            bg="#9E9E9E", fg="white", font=("Arial", 10),
            relief=tk.FLAT, padx=12, cursor="hand2"
        ).pack(side=tk.RIGHT)
        tk.Button(
            bottom_frame, text="📂 開啟 Excel 資料夾", command=self._open_excel_folder,
            bg="#6c757d", fg="white", font=("Arial", 10),
            relief=tk.FLAT, padx=12, cursor="hand2"
        ).pack(side=tk.RIGHT, padx=(0, 6))

    def _section_label(self, parent, text):
        tk.Label(
            parent, text=text,
            font=("Arial", 11, "bold"), bg="#FDFBF6", fg="#5D4037",
            anchor=tk.W
        ).pack(fill=tk.X, pady=(8, 2))

    def _select_json(self):
        path = filedialog.askopenfilename(
            title="選擇自訂句子 JSON",
            filetypes=[("JSON 檔案", "*.json"), ("所有檔案", "*.*")]
        )
        if path:
            self.json_path.set(path)
            self._log(f"📄 JSON：{path}")

    def _select_excel(self):
        path = filedialog.askopenfilename(
            title="選擇原始 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls"), ("所有檔案", "*.*")]
        )
        if path:
            self.excel_path.set(path)
            self._log(f"📊 Excel：{path}")

    def _start_merge(self):
        json_path = self.json_path.get().strip()
        excel_path = self.excel_path.get().strip()

        if not json_path:
            messagebox.showwarning("提示", "請先選擇 JSON 檔案！")
            return
        if not excel_path:
            messagebox.showwarning("提示", "請先選擇 Excel 檔案！")
            return
        if not os.path.exists(json_path):
            messagebox.showerror("錯誤", f"JSON 檔案不存在：\n{json_path}")
            return
        if not os.path.exists(excel_path):
            messagebox.showerror("錯誤", f"Excel 檔案不存在：\n{excel_path}")
            return

        # 確認操作
        confirm = messagebox.askyesno(
            "確認操作",
            f"即將整合並覆蓋原始 Excel 檔案：\n{os.path.basename(excel_path)}\n\n確定繼續？"
        )
        if not confirm:
            return

        self._log("\n" + "=" * 50)
        self._log("🚀 開始整合...")
        self._log("=" * 50)

        # 執行合併
        added, skipped, error = merge_json_into_excel(
            json_path, excel_path,
            progress_callback=self._log
        )

        self._log("\n" + "=" * 50)
        if error:
            self._log(f"❌ 整合失敗：{error}")
            messagebox.showerror("整合失敗", f"發生錯誤：\n{error}")
        else:
            summary = f"整合完成！\n\n新增：{added} 筆\n跳過重複：{skipped} 筆\n\nExcel 已更新並儲存。"
            self._log(f"✅ {summary.replace(chr(10), ' ')}")
            messagebox.showinfo("整合成功", summary)

    def _log(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def _clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _open_excel_folder(self):
        excel_path = self.excel_path.get().strip()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showwarning("提示", "請先選擇 Excel 檔案")
            return
        folder = os.path.dirname(excel_path)
        import platform
        if platform.system() == "Windows":
            os.startfile(folder)
        elif platform.system() == "Darwin":
            os.system(f'open "{folder}"')
        else:
            os.system(f'xdg-open "{folder}"')


# ===== 入口點 =====
if __name__ == "__main__":
    print("=" * 50)
    print("自訂句子整合工具")
    print("=" * 50)
    root = tk.Tk()
    app = MergeApp(root)
    root.mainloop()
    print("程式結束")
