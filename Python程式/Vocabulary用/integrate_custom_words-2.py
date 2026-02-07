#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
整合自訂單字到 Excel 和 JSON
將匯出的 JSON 中的自訂單字整合到 Excel 檔案中,並更新對應的 JSON 檔案
修復版: 支援多種 JSON 格式
修改版: 智慧更新邏輯 - 不清空原有值,衝突時詢問使用者
"""

import json
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class WordIntegrator:
    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()
        self.modified_words = []
        self.new_words = []
        
    def select_file(self, title, filetypes):
        """選擇檔案"""
        filename = filedialog.askopenfilename(
            title=title,
            filetypes=filetypes
        )
        return filename
    
    def load_json_export(self, json_path):
        """載入匯出的 JSON 檔案 - 支援多種格式"""
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 嘗試多種可能的鍵名
            custom_words = None
            found_key = None
            
            # 可能的鍵名列表
            possible_keys = [
                'customWords',           # 新格式
                '自訂單字 (11)',         # 舊格式
                'custom_words',          # 底線格式
                'Custom Words'           # 空格格式
            ]
            
            # 嘗試找到自訂單字
            for key in possible_keys:
                if key in data:
                    custom_words = data[key]
                    found_key = key
                    break
            
            # 如果都沒找到,檢查是否整個 JSON 就是單字字典
            if custom_words is None:
                # 檢查是否有單字物件的特徵
                if isinstance(data, dict) and len(data) > 0:
                    first_value = next(iter(data.values()))
                    if isinstance(first_value, dict) and 'Words' in first_value:
                        custom_words = data
                        found_key = 'root'
            
            if not custom_words:
                messagebox.showwarning(
                    "警告", 
                    f"JSON 檔案中沒有找到自訂單字!\n\n嘗試的鍵名: {', '.join(possible_keys)}\n\nJSON 結構: {list(data.keys())}"
                )
                return None
            
            print(f"✅ 成功載入 JSON (鍵名: {found_key}),找到 {len(custom_words)} 個自訂單字")
            return custom_words
        except Exception as e:
            messagebox.showerror("錯誤", f"載入 JSON 失敗:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def find_word_row(self, ws, word_text):
        """在 Excel 中尋找單字所在的行"""
        # 單字在 F 欄
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'F{row}'].value
            if cell_value and cell_value.strip().lower() == word_text.strip().lower():
                return row
        return None
    
    def copy_row_style(self, source_row, target_row, ws):
        """複製行的樣式"""
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=source_row, column=col)
            target_cell = ws.cell(row=target_row, column=col)
            
            # 複製字體
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    color=source_cell.font.color
                )
            
            # 複製對齊
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=source_cell.alignment.wrap_text
                )
            
            # 複製邊框
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom
                )
            
            # 複製填充
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    fgColor=source_cell.fill.fgColor,
                    bgColor=source_cell.fill.bgColor
                )
    
    def update_cell(self, ws, row, col, new_value, word_text, field_name):
        """智慧更新單一儲存格: 無值不動,有值衝突詢問"""
        if new_value is None or not str(new_value).strip():
            return False  # 無有效新值,不動
        
        current_value = ws[f'{col}{row}'].value
        current_str = str(current_value).strip() if current_value is not None else ''
        
        if current_str:
            # 既有值,詢問是否覆蓋
            msg = f"單字: {word_text}\n欄位: {field_name}\n原值: {current_value}\n新值: {new_value}\n是否覆蓋?"
            if messagebox.askyesno("確認覆蓋", msg):
                ws[f'{col}{row}'] = new_value
                return True
            else:
                return False
        else:
            # 無值,直接覆蓋
            ws[f'{col}{row}'] = new_value
            return True
    
    def update_word_in_excel(self, ws, row, word_obj):
        """更新 Excel 中的單字資料 - 修改版邏輯"""
        updated = False
        
        # 單字 (F 欄): 總是更新 (因為是 key,但允許修正拼寫)
        word_text = word_obj.get('Words', word_obj.get('word', word_obj.get('單字', '')))
        if word_text:
            if self.update_cell(ws, row, 'F', word_text, word_text, 'Words (單字)'):
                updated = True
        
        # 等級 (B 欄)
        level = word_obj.get('等級', None)
        if self.update_cell(ws, row, 'B', level, word_text, '等級'):
            updated = True
        
        # 發音1 (G 欄)
        pron1 = word_obj.get('pronunciation-1', None)
        if self.update_cell(ws, row, 'G', pron1, word_text, 'pronunciation-1'):
            updated = True
        
        # 發音2 (H 欄)
        pron2 = word_obj.get('pronunciation-2', None)
        if self.update_cell(ws, row, 'H', pron2, word_text, 'pronunciation-2'):
            updated = True
        
        # 繁體中文 (I 欄)
        trad_ch = word_obj.get('traditional Chinese', None)
        if self.update_cell(ws, row, 'I', trad_ch, word_text, 'traditional Chinese'):
            updated = True
        
        # 英文解釋 (J 欄)
        eng_mean = word_obj.get('English meaning', None)
        if self.update_cell(ws, row, 'J', eng_mean, word_text, 'English meaning'):
            updated = True
        
        # 分類 (C,D,E 欄): 只更新有提供的,不清空多的
        categories = word_obj.get('分類', None)
        if categories is not None and isinstance(categories, list):
            category_names = ['分類1 (Domain)', '分類2 (Topic)', '分類3 (Source)']
            for i, col in enumerate(['C', 'D', 'E']):
                if i < len(categories):
                    cat_value = categories[i]
                    if self.update_cell(ws, row, col, cat_value, word_text, category_names[i]):
                        updated = True
                # 如果 categories 較短,不動後面的欄位
        
        # 更新 HYPERLINK 公式 (K, L 欄) - 總是基於最新單字更新
        if word_text:
            ws[f'K{row}'].value = f'=HYPERLINK(F{row} & " - sentence.mp3", F{row})'
            ws[f'L{row}'].value = f'=HYPERLINK(F{row} & ".mp3", F{row})'
            updated = True  # 公式更新算更新
        
        return updated
    
    def add_new_word_to_excel(self, ws, word_obj):
        """在 Excel 中新增單字 - 直接寫入 JSON 值"""
        new_row = ws.max_row + 1
        
        # 複製前一行的樣式
        if new_row > 2:
            self.copy_row_style(new_row - 1, new_row, ws)
        
        # 寫入資料 (新增時無需詢問,直接寫)
        word_text = word_obj.get('Words', word_obj.get('word', word_obj.get('單字', '')))
        ws[f'F{new_row}'] = word_text
        ws[f'B{new_row}'] = word_obj.get('等級', '未分類')
        ws[f'G{new_row}'] = word_obj.get('pronunciation-1', '')
        ws[f'H{new_row}'] = word_obj.get('pronunciation-2', '')
        ws[f'I{new_row}'] = word_obj.get('traditional Chinese', '')
        ws[f'J{new_row}'] = word_obj.get('English meaning', '')
        
        # 分類
        categories = word_obj.get('分類', [])
        ws[f'C{new_row}'] = categories[0] if len(categories) > 0 else ''
        ws[f'D{new_row}'] = categories[1] if len(categories) > 1 else ''
        ws[f'E{new_row}'] = categories[2] if len(categories) > 2 else ''
        
        # HYPERLINK
        if word_text:
            ws[f'K{new_row}'].value = f'=HYPERLINK(F{new_row} & " - sentence.mp3", F{new_row})'
            ws[f'L{new_row}'].value = f'=HYPERLINK(F{new_row} & ".mp3", F{new_row})'
        
        return new_row
    
    def update_excel(self, excel_path, custom_words):
        """更新 Excel 檔案"""
        try:
            print(f"📂 正在載入 Excel: {excel_path}")
            wb = load_workbook(excel_path)
            ws = wb.active
            
            for word_text, word_obj in custom_words.items():
                # 確保 word_obj 有 Words 欄位
                if 'Words' not in word_obj:
                    word_obj['Words'] = word_text
                
                # 尋找單字是否已存在
                existing_row = self.find_word_row(ws, word_text)
                
                if existing_row:
                    # 更新現有單字
                    print(f"✏️  更新單字: {word_text} (第 {existing_row} 行)")
                    if self.update_word_in_excel(ws, existing_row, word_obj):
                        self.modified_words.append(word_text)
                else:
                    # 新增單字
                    print(f"➕ 新增單字: {word_text}")
                    new_row = self.add_new_word_to_excel(ws, word_obj)
                    self.new_words.append(word_text)
            
            # 儲存 Excel
            wb.save(excel_path)
            print(f"💾 Excel 已儲存: {excel_path}")
            
            return True
        except Exception as e:
            messagebox.showerror("錯誤", f"更新 Excel 失敗:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def excel_to_json(self, excel_path):
        """將 Excel 轉換為 JSON"""
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            
            words_list = []
            
            # 從第二行開始讀取 (第一行是標題)
            for row in range(2, ws.max_row + 1):
                word_obj = {
                    '等級': ws[f'B{row}'].value or '未分類',
                    'Words': ws[f'F{row}'].value or '',
                    'pronunciation-1': ws[f'G{row}'].value or '',
                    'pronunciation-2': ws[f'H{row}'].value or '',
                    'traditional Chinese': ws[f'I{row}'].value or '',
                    'English meaning': ws[f'J{row}'].value or '',
                    'Unnamed: 10': ws[f'F{row}'].value or '',  # 複製單字
                    'Unnamed: 11': ws[f'F{row}'].value or '',  # 複製單字
                    '分類': []
                }
                
                # 讀取分類
                domain = ws[f'C{row}'].value
                topic = ws[f'D{row}'].value
                source = ws[f'E{row}'].value
                
                categories = []
                if domain:
                    categories.append(domain)
                if topic:
                    categories.append(topic)
                if source:
                    categories.append(source)
                
                word_obj['分類'] = categories
                
                # 只添加有單字的行
                if word_obj['Words']:
                    words_list.append(word_obj)
            
            # 建立 JSON 結構
            json_data = {
                "New Words": words_list
            }
            
            return json_data
        except Exception as e:
            messagebox.showerror("錯誤", f"轉換 JSON 失敗:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return None
    
    def save_json(self, json_data, output_path):
        """儲存 JSON 檔案"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            print(f"💾 JSON 已儲存: {output_path}")
            return True
        except Exception as e:
            messagebox.showerror("錯誤", f"儲存 JSON 失敗:\n{str(e)}")
            import traceback
            traceback.print_exc()
            return False
    
    def show_summary(self):
        """顯示更新摘要"""
        summary = "=" * 50 + "\n"
        summary += "✅ 整合完成!\n"
        summary += "=" * 50 + "\n\n"
        
        if self.modified_words:
            summary += f"📝 修改的單字 ({len(self.modified_words)} 個):\n"
            for word in self.modified_words:
                summary += f"  - {word}\n"
            summary += "\n"
        
        if self.new_words:
            summary += f"➕ 新增的單字 ({len(self.new_words)} 個):\n"
            for word in self.new_words:
                summary += f"  - {word}\n"
            summary += "\n"
        
        if not self.modified_words and not self.new_words:
            summary += "ℹ️  沒有任何變更\n\n"
        
        summary += "=" * 50 + "\n"
        
        print(summary)
        messagebox.showinfo("整合完成", summary)
    
    def run(self):
        """執行主流程"""
        print("=" * 60)
        print("🚀 單字整合工具 (修復版 - 智慧更新)")
        print("=" * 60)
        
        # 1. 選擇匯出的 JSON 檔案
        print("\n步驟 1: 請選擇匯出的 JSON 檔案...")
        json_path = self.select_file(
            "選擇匯出的 JSON 檔案",
            [("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if not json_path:
            print("❌ 未選擇 JSON 檔案,程式結束")
            return
        
        print(f"✅ 已選擇: {json_path}")
        
        # 2. 載入自訂單字
        custom_words = self.load_json_export(json_path)
        if not custom_words:
            return
        
        # 3. 選擇要更新的 Excel 檔案
        print("\n步驟 2: 請選擇要更新的 Excel 檔案...")
        excel_path = self.select_file(
            "選擇要更新的 Excel 檔案",
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not excel_path:
            print("❌ 未選擇 Excel 檔案,程式結束")
            return
        
        print(f"✅ 已選擇: {excel_path}")
        
        # 4. 更新 Excel
        print("\n步驟 3: 正在更新 Excel...")
        if not self.update_excel(excel_path, custom_words):
            return
        
        # 5. 轉換為 JSON
        print("\n步驟 4: 正在轉換為 JSON...")
        json_data = self.excel_to_json(excel_path)
        if not json_data:
            return
        
        # 6. 儲存 JSON (與 Excel 同路徑同檔名)
        excel_dir = os.path.dirname(excel_path)
        excel_basename = os.path.splitext(os.path.basename(excel_path))[0]
        json_output_path = os.path.join(excel_dir, f"{excel_basename}.json")
        
        print(f"\n步驟 5: 正在儲存 JSON 到 {json_output_path}...")
        if not self.save_json(json_data, json_output_path):
            return
        
        # 7. 顯示摘要
        self.show_summary()


if __name__ == "__main__":
    integrator = WordIntegrator()
    integrator.run()