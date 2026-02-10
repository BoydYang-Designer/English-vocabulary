import pandas as pd
import openpyxl
import re
import json
import os
import sys
from tkinter import Tk, filedialog, messagebox

def select_files():
    """使用 GUI 選擇檔案 A 和檔案 B"""
    root = Tk()
    root.withdraw()  # 隱藏主視窗
    
    messagebox.showinfo("選擇檔案", "請選擇檔案 A (Z_total_words.xlsx 或來源檔案)")
    excel_a_path = filedialog.askopenfilename(
        title="選擇檔案 A (來源檔案)",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    if not excel_a_path:
        messagebox.showerror("錯誤", "未選擇檔案 A，程式結束")
        sys.exit(1)
    
    messagebox.showinfo("選擇檔案", "請選擇檔案 B (sentence.xlsx 或目標檔案)")
    excel_b_path = filedialog.askopenfilename(
        title="選擇檔案 B (目標檔案)",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    if not excel_b_path:
        messagebox.showerror("錯誤", "未選擇檔案 B，程式結束")
        sys.exit(1)
    
    root.destroy()
    return excel_a_path, excel_b_path

def extract_eg_sentences(text):
    """從 'English meaning' 欄位中提取以 'E.g.' 開頭的句子"""
    if pd.isna(text):
        return []
    sentences = re.findall(r'E\.g\.\s*(.*?)(?:\.\s*$|\.\s+|\n|$)', text, re.MULTILINE)
    cleaned_sentences = [s.strip() for s in sentences if s.strip()]
    print(f"提取的例句: {cleaned_sentences}")
    return cleaned_sentences

def get_max_suffix(word, df_b):
    """獲取 Excel B 中該單字的最大編號"""
    pattern = rf'^{re.escape(word)}-(\d+)$'
    suffixes = df_b['Words'].str.extract(pattern).dropna()
    if suffixes.empty:
        return 0
    try:
        suffixes = suffixes.astype(int)
        return int(suffixes.max().item())
    except Exception as e:
        print(f"處理單字 {word} 的編號時發生錯誤: {e}")
        print(f"問題數據: {suffixes}")
        raise

def preprocess_sentence(sentence):
    """預處理句子：去除標點符號並轉換為小寫"""
    sentence = re.sub(r'[^\w\s]', '', sentence)
    return sentence.lower().strip()

def update_excel_b(excel_a_path, excel_b_path, output_path):
    """將 Excel A 的新例句更新到 Excel B，並更新現有記錄的分類和等級"""
    if not os.path.exists(excel_a_path):
        raise FileNotFoundError(f"找不到 Excel A 檔案: {excel_a_path}")
    if not os.path.exists(excel_b_path):
        raise FileNotFoundError(f"找不到 Excel B 檔案: {excel_b_path}")
    
    # 獲取原始檔案大小
    input_size = os.path.getsize(excel_b_path) / 1024  # 單位：KB
    print(f"Excel B 原始檔案大小: {input_size:.2f} KB")
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    print(f"Excel A 行數: {len(df_a)}")
    print(f"Excel B 原始行數: {len(df_b)}")
    print(f"Excel A 欄位: {list(df_a.columns)}")
    print(f"Excel B 欄位: {list(df_b.columns)}")
    
    wb = openpyxl.load_workbook(excel_b_path)
    ws = wb.active
    
    # 確認 Excel B 的欄位索引
    b_columns = {col: idx for idx, col in enumerate(df_b.columns, 1)}
    print(f"Excel B 欄位對應: {b_columns}")
    
    new_rows = []
    
    for index, row in df_a.iterrows():
        word = row['Words']
        if pd.isna(word) or not isinstance(word, str):
            print(f"跳過無效單字: {word}")
            continue
        
        eg_sentences = extract_eg_sentences(row.get('English meaning', ''))

        # 🔹 從檔案 A 讀取對應欄位
        level_a = str(row['等級']).strip() if pd.notna(row.get('等級')) else ''
        category1 = str(row['分類1']).strip() if pd.notna(row.get('分類1')) else ''
        category2 = str(row['分類2']).strip() if pd.notna(row.get('分類2')) else ''
        category3 = str(row['分類3']).strip() if pd.notna(row.get('分類3')) else ''
        
        print(f"處理單字: {word}, 等級: {level_a}, 分類1: {category1}, 分類2: {category2}, 分類3: {category3}")
        
        word_records = df_b[df_b['Words'].notna() & df_b['Words'].str.match(rf'^{re.escape(word)}-\d+$')]
        max_suffix = get_max_suffix(word, df_b)
        
        existing_sentences = set(preprocess_sentence(s) for s in word_records['句子'].dropna())
        
        # 更新現有記錄的分類和等級
        for b_index, b_row in word_records.iterrows():
            b_row_idx = b_index + 2  # Excel 行號 (從 1 開始，加上表頭)
            update_needed = False
            
            # 更新等級（如果 B 的等級為空且 A 有等級）
            current_level = str(b_row['等級']).strip() if pd.notna(b_row.get('等級')) else ''
            if not current_level and level_a:
                ws.cell(row=b_row_idx, column=b_columns['等級'], value=level_a)
                update_needed = True
                print(f"  更新等級: 行 {b_row_idx}, {current_level} → {level_a}")
            
            # 更新分類1
            if category1:
                ws.cell(row=b_row_idx, column=b_columns['分類1'], value=category1)
                update_needed = True
            
            # 更新分類2
            if category2:
                ws.cell(row=b_row_idx, column=b_columns['分類2'], value=category2)
                update_needed = True
            
            # 更新分類3
            if category3:
                ws.cell(row=b_row_idx, column=b_columns['分類3'], value=category3)
                update_needed = True
                
            if update_needed:
                print(f"  更新行 {b_row_idx}: Words={b_row['Words']}, 等級={level_a}, 分類=({category1}, {category2}, {category3})")
        
        # 新增例句
        if eg_sentences:
            for sentence in eg_sentences:
                preprocessed_sentence = preprocess_sentence(sentence)
                if preprocessed_sentence not in existing_sentences:
                    max_suffix += 1
                    new_word = f"{word}-{max_suffix}"
                    new_row = {
                        'Unnamed: 0': '',
                        '音檔': '',
                        '等級': level_a,
                        '分類1': category1,
                        '分類2': category2,
                        '分類3': category3,
                        'Words': new_word,
                        '名人': '',
                        '句子': sentence,
                        '中文': '',
                        '記錄': ''
                    }
                    new_rows.append(new_row)
                    existing_sentences.add(preprocessed_sentence)
                    print(f"  添加新例句: {new_word}, 句子={sentence}")
    
    # 將新行添加到工作表
    if new_rows:
        for row_dict in new_rows:
            row_values = [row_dict.get(col, '') for col in df_b.columns]
            ws.append(row_values)
    
    print(f"\n新增行數: {len(new_rows)}")
    
    wb.save(output_path)
    
    df_output = pd.read_excel(output_path)
    output_size = os.path.getsize(output_path) / 1024  # 單位：KB
    print(f"Excel B 輸出行數: {len(df_output)}")
    print(f"Excel B 輸出檔案大小: {output_size:.2f} KB")
    
    if len(df_output) < len(df_b):
        print("警告：輸出行數少於原始行數，可能有數據丟失！")
    
    print(f"已更新 Excel B，結果保存到 {output_path}")
    return True

def compare_excel_files(excel_a_path, excel_b_path, output_json_path):
    """比對兩個 Excel 檔案的指定欄位（不含音檔），並在有差異時記錄到 JSON 檔案"""
    columns_to_compare = ['等級', '分類1', '分類2', '分類3', 'Words', '名人', '句子', '中文']
    
    if not os.path.exists(excel_a_path):
        raise FileNotFoundError(f"找不到 Excel A 檔案: {excel_a_path}")
    if not os.path.exists(excel_b_path):
        raise FileNotFoundError(f"找不到 Excel B 檔案: {excel_b_path}")
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    for col in columns_to_compare:
        if col not in df_a.columns:
            df_a[col] = ''
        if col not in df_b.columns:
            df_b[col] = ''
    
    df_a = df_a.fillna('')
    df_b = df_b.fillna('')
    
    a_dict = {row['Words']: row.to_dict() for _, row in df_a.iterrows() if pd.notna(row['Words'])}
    b_dict = {row['Words']: row.to_dict() for _, row in df_b.iterrows() if pd.notna(row['Words'])}
    
    differences = []
    
    common_words = set(a_dict.keys()) & set(b_dict.keys())
    for word in common_words:
        a_row = a_dict[word]
        b_row = b_dict[word]
        diff = {}
        
        for col in columns_to_compare:
            if col != 'Words':
                a_val = a_row.get(col, '')
                b_val = b_row.get(col, '')
                if a_val != b_val:
                    diff[col] = {'A 值': a_val, 'B 值': b_val}
        
        if diff:
            differences.append({
                'Words': word,
                '狀態': 'A 和 B 均有，但內容有差異',
                '差異內容': diff
            })
    
    a_only_words = set(a_dict.keys()) - set(b_dict.keys())
    for word in a_only_words:
        differences.append({
            'Words': word,
            '狀態': 'A 有，B 無',
            'A 內容': {k: v for k, v in a_dict[word].items() if k in columns_to_compare}
        })
    
    b_only_words = set(b_dict.keys()) - set(a_dict.keys())
    for word in b_only_words:
        differences.append({
            'Words': word,
            '狀態': 'B 有，A 無',
            'B 內容': {k: v for k, v in b_dict[word].items() if k in columns_to_compare}
        })
    
    if differences:
        with open(output_json_path, 'w', encoding='utf-8') as json_file:
            json.dump(differences, json_file, ensure_ascii=False, indent=4)
        print(f"\n比對完成，差異已記錄到 {output_json_path}")
        print(f"共發現 {len(differences)} 項差異")
    else:
        print("\nA 和 B 之間沒有差異，未生成 JSON 檔案")

def main():
    """主函數，使用 GUI 選擇檔案，更新 Excel B，然後比對結果"""
    try:
        # 使用 GUI 選擇檔案
        excel_a_path, excel_b_path = select_files()
        
        print(f"檔案 A: {excel_a_path}")
        print(f"檔案 B: {excel_b_path}")
        
        # 生成輸出檔案路徑
        b_dir = os.path.dirname(excel_b_path)
        b_name = os.path.basename(excel_b_path)
        b_name_without_ext = os.path.splitext(b_name)[0]
        
        output_path = os.path.join(b_dir, f'updated_{b_name}')
        output_json_path = os.path.join(b_dir, 'comparison_result.json')
        
        print(f"輸出檔案: {output_path}")
        print(f"比對結果: {output_json_path}")
        print("\n" + "="*50)
        
        # 驗證檔案 B 的 Words 欄位格式
        df_b = pd.read_excel(excel_b_path)
        invalid_words = df_b[df_b['Words'].notna() & ~df_b['Words'].str.match(r'^[\w\s\''éèêëáàâãäåíìîïóòôõöúùûüýÿ-]+-\d+$', na=False)]
        if not invalid_words.empty:
            print("發現無效的 Words 欄位值：")
            print(invalid_words[['Words']])
            messagebox.showwarning("警告", "檔案 B 中有無效的 Words 欄位格式，請檢查")
        
        # 更新 Excel B
        update_excel_b(excel_a_path, excel_b_path, output_path)
        
        # 比對原始 B 檔案和更新後的檔案
        compare_excel_files(excel_b_path, output_path, output_json_path)
        
        messagebox.showinfo("完成", f"更新完成！\n\n輸出檔案: {output_path}\n比對結果: {output_json_path}")
        
    except FileNotFoundError as e:
        print(f"錯誤: {e}")
        messagebox.showerror("錯誤", str(e))
        sys.exit(1)
    except Exception as e:
        print(f"發生未知錯誤: {e}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("錯誤", f"發生未知錯誤: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
