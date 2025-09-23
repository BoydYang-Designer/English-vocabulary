import pandas as pd
import os
from openai import OpenAI
from tkinter import Tk, filedialog
from openpyxl import load_workbook

# 初始化 OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# 選擇 Excel 檔
Tk().withdraw()
file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
df = pd.read_excel(file_path)

# 檢查中文欄位空值
print(f"中文欄位空值的行數: {df['中文'].isna().sum()}")

# 使用者輸入要生成的數量
num_to_generate = int(input("請輸入要生成的數量（例如 10）："))

# 篩選中文欄位空值的列
df_to_process = df[df['中文'].isna()].head(num_to_generate)
print(f"待處理資料: \n{df_to_process[['句子', 'Words', '中文']]}")

# Excel 資料夾路徑
folder_path = os.path.dirname(file_path)

# 儲存翻譯結果
for index, row in df_to_process.iterrows():
    sentence = row['句子']
    word = row['Words']
    
    # 生成 mp3 音檔
    mp3_path = os.path.join(folder_path, f"{word}.mp3")
    with client.audio.speech.with_streaming_response.create(
        model="gpt-4o-mini-tts",
        voice="alloy",
        input=sentence
    ) as response:
        response.stream_to_file(mp3_path)
    
    # 翻譯成繁體中文
    translation_resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "請將英文翻譯成繁體中文。"},
            {"role": "user", "content": sentence}
        ]
    )
    translated_text = translation_resp.choices[0].message.content
    print(f"翻譯結果 for {sentence}: {translated_text}")
    df.loc[index, '中文'] = translated_text  # 更新 DataFrame
    print(f"更新 DataFrame 索引 {index}，中文: {df.loc[index, '中文']}")  # 確認 DataFrame 更新

# 使用 openpyxl 更新 Excel，僅修改中文欄位
book = load_workbook(file_path)
sheet = book.active
header = [cell.value for cell in sheet[1]]
print(f"Excel 標題: {header}")

# 找到「中文」欄位的列索引
try:
    chinese_col = header.index('中文') + 1
except ValueError:
    print("錯誤：Excel 標題中未找到 '中文' 欄位，請檢查欄位名稱")
    exit()

# 更新「中文」欄位，直接從 df 讀取
for index in df_to_process.index:
    excel_row = index + 2  # 假設第一行為標題，資料從第二行開始
    translated_text = df.loc[index, '中文']  # 從更新後的 DataFrame 取得中文
    print(f"更新 Excel 行 {excel_row}，中文: {translated_text}")
    sheet.cell(row=excel_row, column=chinese_col).value = translated_text

# 儲存 Excel 檔案
print(f"儲存檔案到: {file_path}")
book.save(file_path)
print(f"✅ 完成 {len(df_to_process)} 個 Words 的音檔與中文翻譯")