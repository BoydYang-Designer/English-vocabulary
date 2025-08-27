import pandas as pd
import openpyxl
import re
import json
import os
import sys

def extract_eg_sentences(text):
    """å¾ 'English meaning' æ¬„ä½ä¸­æå–ä»¥ 'E.g.' é–‹é ­çš„å¥å­ï¼Œè™•ç†ç©ºæ ¼æ•¸é‡èª¤æ¤"""
    if pd.isna(text):
        return []
    sentences = re.findall(r'E\.g\.\s*(.*?)(?:\.\s*$|\.\s+|\n|$)', text, re.MULTILINE)
    cleaned_sentences = [s.strip() for s in sentences if s.strip()]
    print(f"æå–çš„ä¾‹å¥: {cleaned_sentences}")
    return cleaned_sentences

def get_max_suffix(word, df_b):
    """ç²å– Excel B ä¸­è©²å–®å­—çš„æœ€å¤§ç·¨è™Ÿ"""
    pattern = rf'^{re.escape(word)}-(\d+)$'
    suffixes = df_b['Words'].str.extract(pattern).dropna()
    if suffixes.empty:
        return 0
    try:
        suffixes = suffixes.astype(int)
        return int(suffixes.max().item())
    except Exception as e:
        print(f"è™•ç†å–®å­— {word} çš„ç·¨è™Ÿæ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        print(f"å•é¡Œæ•¸æ“š: {suffixes}")
        raise

def preprocess_sentence(sentence):
    """é è™•ç†å¥å­ï¼šå»é™¤æ¨™é»ç¬¦è™Ÿä¸¦è½‰æ›ç‚ºå°å¯«"""
    sentence = re.sub(r'[^\w\s]', '', sentence)
    return sentence.lower().strip()

def update_excel_b(excel_a_path, excel_b_path, output_path):
    """å°‡ Excel A çš„æ–°ä¾‹å¥æ›´æ–°åˆ° Excel Bï¼Œä¸¦æ›´æ–°ç¾æœ‰è¨˜éŒ„çš„åˆ†é¡å’Œç­‰ç´š"""
    if not os.path.exists(excel_a_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ° Excel A æª”æ¡ˆ: {excel_a_path}")
    if not os.path.exists(excel_b_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ° Excel B æª”æ¡ˆ: {excel_b_path}")
    
    # ç²å–åŸå§‹æª”æ¡ˆå¤§å°
    input_size = os.path.getsize(excel_b_path) / 1024  # å–®ä½ï¼šKB
    print(f"Excel B åŸå§‹æª”æ¡ˆå¤§å°: {input_size:.2f} KB")
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    print(f"Excel A è¡Œæ•¸: {len(df_a)}")
    print(f"Excel B åŸå§‹è¡Œæ•¸: {len(df_b)}")
    
    wb = openpyxl.load_workbook(excel_b_path)
    ws = wb.active
    
    new_rows = []
    
    for index, row in df_a.iterrows():
        word = row['Words']
        if pd.isna(word) or not isinstance(word, str):
            print(f"è·³éç„¡æ•ˆå–®å­—: {word}")
            continue
        
        eg_sentences = extract_eg_sentences(row['English meaning'])

        # ğŸ”¹ æ”¯æ´å¤šåˆ†é¡
        categories = []
        for c in ['åˆ†é¡1', 'åˆ†é¡2', 'åˆ†é¡3']:
            if c in row and pd.notna(row[c]):
                categories.append(str(row[c]).strip())
        category1 = categories[0] if len(categories) > 0 else ''
        category2 = categories[1] if len(categories) > 1 else ''
        category3 = categories[2] if len(categories) > 2 else ''

        level_a = row['ç­‰ç´š'] if pd.notna(row['ç­‰ç´š']) else ''
        
        print(f"è™•ç†å–®å­—: {word}, åˆ†é¡: {categories}, ç­‰ç´š: {level_a}")
        
        word_records = df_b[df_b['Words'].notna() & df_b['Words'].str.match(rf'^{re.escape(word)}-\d+$')]
        max_suffix = get_max_suffix(word, df_b)
        
        existing_sentences = set(preprocess_sentence(s) for s in word_records['å¥å­'].dropna())
        
        for b_index, b_row in word_records.iterrows():
            b_row_idx = b_index + 2
            update_needed = False
            
            # æ›´æ–°åˆ†é¡
            if category1 or category2 or category3:
                ws.cell(row=b_row_idx, column=3, value=category1)
                ws.cell(row=b_row_idx, column=4, value=category2)
                ws.cell(row=b_row_idx, column=5, value=category3)
                update_needed = True
            
            current_level = b_row['ç­‰ç´š'] if pd.notna(b_row['ç­‰ç´š']) else ''
            if not current_level and level_a:
                ws.cell(row=b_row_idx, column=2, value=level_a)
                update_needed = True
                
            if update_needed:
                print(f"æ›´æ–°è¡Œ {b_row_idx}: Words={b_row['Words']}, åˆ†é¡=({category1},{category2},{category3}), ç­‰ç´š={level_a}")
        
        if eg_sentences:
            for sentence in eg_sentences:
                preprocessed_sentence = preprocess_sentence(sentence)
                if preprocessed_sentence not in existing_sentences:
                    max_suffix += 1
                    new_word = f"{word}-{max_suffix}"
                    new_row = {
                        'éŸ³æª”': '',
                        'ç­‰ç´š': level_a,
                        'åˆ†é¡1': category1,
                        'åˆ†é¡2': category2,
                        'åˆ†é¡3': category3,
                        'Words': new_word,
                        'åäºº': '',
                        'å¥å­': sentence,
                        'ä¸­æ–‡': ''
                    }
                    new_rows.append(new_row)
                    existing_sentences.add(preprocessed_sentence)
                    print(f"æ·»åŠ æ–°ä¾‹å¥: {new_word}, å¥å­={sentence}")
    
    if new_rows:
        for row in pd.DataFrame(new_rows).itertuples(index=False):
            ws.append(row)
    
    print(f"æ–°å¢è¡Œæ•¸: {len(new_rows)}")
    
    wb.save(output_path)
    
    df_output = pd.read_excel(output_path)
    output_size = os.path.getsize(output_path) / 1024  # å–®ä½ï¼šKB
    print(f"Excel B è¼¸å‡ºè¡Œæ•¸: {len(df_output)}")
    print(f"Excel B è¼¸å‡ºæª”æ¡ˆå¤§å°: {output_size:.2f} KB")
    
    if len(df_output) < len(df_b):
        print("è­¦å‘Šï¼šè¼¸å‡ºè¡Œæ•¸å°‘æ–¼åŸå§‹è¡Œæ•¸ï¼Œå¯èƒ½æœ‰æ•¸æ“šä¸Ÿå¤±ï¼")
    
    print(f"å·²æ›´æ–° Excel Bï¼Œçµæœä¿å­˜åˆ° {output_path}")
    return True

def compare_excel_files(excel_a_path, excel_b_path, output_json_path):
    """æ¯”å°å…©å€‹ Excel æª”æ¡ˆçš„æŒ‡å®šæ¬„ä½ï¼ˆä¸å«éŸ³æª”ï¼‰ï¼Œä¸¦åœ¨æœ‰å·®ç•°æ™‚è¨˜éŒ„åˆ° JSON æª”æ¡ˆ"""
    columns_to_compare = ['ç­‰ç´š', 'åˆ†é¡1', 'åˆ†é¡2', 'åˆ†é¡3', 'Words', 'åäºº', 'å¥å­', 'ä¸­æ–‡']
    
    if not os.path.exists(excel_a_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ° Excel A æª”æ¡ˆ: {excel_a_path}")
    if not os.path.exists(excel_b_path):
        raise FileNotFoundError(f"æ‰¾ä¸åˆ° Excel B æª”æ¡ˆ: {excel_b_path}")
    
    df_a = pd.read_excel(excel_a_path)
    df_b = pd.read_excel(excel_b_path)
    
    for col in columns_to_compare:
        if col not in df_a.columns:
            df_a[col] = ''
        if col not in df_b.columns:
            df_b[col] = ''
    
    df_a = df_a.fillna('')
    df_b = df_b.fillna('')
    
    a_dict = {row['Words']: row.to_dict() for _, row in df_a.iterrows()}
    b_dict = {row['Words']: row.to_dict() for _, row in df_b.iterrows()}
    
    differences = []
    
    common_words = set(a_dict.keys()) & set(b_dict.keys())
    for word in common_words:
        a_row = a_dict[word]
        b_row = b_dict[word]
        diff = {}
        
        for col in columns_to_compare:
            if col != 'Words':
                a_val = a_row[col]
                b_val = b_row[col]
                if a_val != b_val:
                    diff[col] = {'A å€¼': a_val, 'B å€¼': b_val}
        
        if diff:
            differences.append({
                'Words': word,
                'ç‹€æ…‹': 'A å’Œ B å‡æœ‰ï¼Œä½†å…§å®¹æœ‰å·®ç•°',
                'å·®ç•°å…§å®¹': diff
            })
    
    a_only_words = set(a_dict.keys()) - set(b_dict.keys())
    for word in a_only_words:
        differences.append({
            'Words': word,
            'ç‹€æ…‹': 'A æœ‰ï¼ŒB ç„¡',
            'A å…§å®¹': {k: v for k, v in a_dict[word].items() if k in columns_to_compare}
        })
    
    b_only_words = set(b_dict.keys()) - set(a_dict.keys())
    for word in b_only_words:
        differences.append({
            'Words': word,
            'ç‹€æ…‹': 'B æœ‰ï¼ŒA ç„¡',
            'B å…§å®¹': {k: v for k, v in b_dict[word].items() if k in columns_to_compare}
        })
    
    if differences:
        with open(output_json_path, 'w', encoding='utf-8') as json_file:
            json.dump(differences, json_file, ensure_ascii=False, indent=4)
        print(f"æ¯”å°å®Œæˆï¼Œå·®ç•°å·²è¨˜éŒ„åˆ° {output_json_path}")
    else:
        print("A å’Œ B ä¹‹é–“æ²’æœ‰å·®ç•°ï¼Œæœªç”Ÿæˆ JSON æª”æ¡ˆ")

def main():
    """ä¸»å‡½æ•¸ï¼Œå…ˆæ›´æ–° Excel Bï¼Œç„¶å¾Œæ¯”å° sentence.xlsx å’Œ updated_sentence.xlsx"""
    excel_a_path = 'Z_total_words.xlsx'
    excel_b_path = 'sentence.xlsx'
    output_path = 'updated_sentence.xlsx'
    output_json_path = 'comparison_result.json'
    
    try:
        df_b = pd.read_excel(excel_b_path)
        # å…è¨±å­—æ¯ã€æ•¸å­—ã€é€£å­—ç¬¦ã€ç©ºæ ¼ã€æ’‡è™Ÿå’Œé‡éŸ³ç¬¦è™Ÿ
        invalid_words = df_b[df_b['Words'].notna() & ~df_b['Words'].str.match(r'^[\w\s\'â€™Ã©Ã¨ÃªÃ«Ã¡Ã Ã¢Ã£Ã¤Ã¥Ã­Ã¬Ã®Ã¯Ã³Ã²Ã´ÃµÃ¶ÃºÃ¹Ã»Ã¼Ã½Ã¿-]+-\d+$', na=False)]
        if not invalid_words.empty:
            print("ç™¼ç¾ç„¡æ•ˆçš„ Words æ¬„ä½å€¼ï¼š")
            print(invalid_words[['Words']])
            return
    except Exception as e:
        print(f"è®€å– Excel B æ™‚ç™¼ç”ŸéŒ¯èª¤: {e}")
        return
    
    try:
        update_excel_b(excel_a_path, excel_b_path, output_path)
        compare_excel_files(excel_b_path, output_path, output_json_path)
    except FileNotFoundError as e:
        print(f"éŒ¯èª¤: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"ç™¼ç”ŸæœªçŸ¥éŒ¯èª¤: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
