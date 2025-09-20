import openpyxl
import requests
from bs4 import BeautifulSoup
import time
import tkinter as tk
from tkinter import filedialog

def get_cefr_level(word):
    """
    從劍橋詞典網站抓取單字的CEFR等級。
    """
    if not word or not isinstance(word, str):
        return None

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    url = f"https://dictionary.cambridge.org/dictionary/english/{word.lower()}"
    
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        level_tag = soup.find('span', class_='epp-xref')
        
        if level_tag:
            return level_tag.text.strip()
        else:
            return None
            
    except requests.RequestException as e:
        print(f"抓取 '{word}' 時發生網路錯誤: {e}")
        return None

def update_excel_levels(filename):
    """
    讀取指定的Excel檔案，抓取單字等級並更新檔案。
    """
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        
        header = [cell.value for cell in sheet[1]]
        try:
            level_col_idx = header.index('等級') + 1
            word_col_idx = header.index('Words') + 1
        except ValueError:
            print("錯誤：找不到 '等級' 或 'Words' 欄位。請檢查 Excel 標頭。")
            return

        print("開始處理檔案...")
        for row_num in range(2, sheet.max_row + 1):
            level_cell = sheet.cell(row=row_num, column=level_col_idx)
            word_cell = sheet.cell(row=row_num, column=word_col_idx)
            
            if not level_cell.value and word_cell.value:
                word = word_cell.value
                print(f"正在查詢單字: '{word}'...")
                
                level = get_cefr_level(word)
                
                if level:
                    level_cell.value = level
                    print(f"  -> 找到等級 '{level}'，已填入。")
                else:
                    print(f"  -> 未找到 '{word}' 的等級。")
                
                time.sleep(1)

        workbook.save(filename)
        print(f"\n處理完成！檔案 '{filename}' 已更新。")

    except FileNotFoundError:
        print(f"錯誤：找不到檔案 '{filename}'。")
    except Exception as e:
        print(f"處理過程中發生未預期的錯誤: {e}")

def select_file_and_run():
    """
    彈出檔案選擇視窗，並執行更新流程。
    """
    # 建立一個隱藏的 tkinter 根視窗
    root = tk.Tk()
    root.withdraw()

    # 打開檔案選擇對話框，限制只能選擇 .xlsx 檔案
    file_path = filedialog.askopenfilename(
        title="請選擇要更新的 Excel 檔案",
        filetypes=(("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*"))
    )

    if file_path:
        print(f"已選擇檔案: {file_path}")
        update_excel_levels(filename=file_path)
    else:
        print("未選擇任何檔案，程式已結束。")

# --- 執行主程式 ---
if __name__ == "__main__":
    select_file_and_run()