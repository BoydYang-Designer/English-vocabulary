import os
import re
import json
import requests
import tkinter as tk
from tkinter import filedialog, simpledialog
import openpyxl
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup

def fetch_yahoo(word, save_folder="."):
    """
    抓取指定單字的 KK 音標和語音檔。
    """
    result = {
        "kk": None,
        "audio_file": None,
        "status": "失敗", # 預設狀態為失敗
        "error": None
    }

    # --- 1. 下載音檔 ---
    mp3_url = f"https://s.yimg.com/bg/dict/dreye/live/f/{word}.mp3"
    mp3_path = os.path.join(save_folder, f"{word}.mp3")
    audio_success = False
    try:
        r = requests.get(mp3_url, timeout=10)
        if r.status_code == 200 and r.content:
            with open(mp3_path, "wb") as f:
                f.write(r.content)
            result["audio_file"] = mp3_path
            audio_success = True
        else:
             print(f"  - 音檔下載失敗 (HTTP 狀態碼: {r.status_code})")
    except Exception as e:
        print(f"  - 音檔下載失敗: {e}")

    # --- 2. 網頁抓取 KK 音標 ---
    url = f"https://tw.dictionary.search.yahoo.com/search?p={word}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"}
    kk_success = False
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(resp.text, "html.parser")

        if soup.find("h3", string=re.compile("找不到")):
            result["error"] = "在字典中找不到該單字"
        else:
            # 抓音標 [gʊd]
            kk_match = soup.find(string=re.compile(r"KK\[[^\]]+\]"))
            if kk_match:
                m = re.search(r"\[([^\]]+)\]", kk_match)
                if m:
                    result["kk"] = f"[{m.group(1)}]"
                    kk_success = True
            
            if not kk_success:
                 result["error"] = "無法在頁面中定位到 KK 音標"

    except Exception as e:
        result["error"] = f"網路請求失敗: {e}"

    # --- 3. 判斷最終狀態 ---
    # 必須音標和音檔都成功，才算成功
    if kk_success and audio_success:
        result["status"] = "成功"

    return result

def main():
    root = tk.Tk()
    root.withdraw()

    excel_path = filedialog.askopenfilename(title="選擇 Excel 檔案", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        print("未選擇檔案，程式結束。")
        return

    # ▼▼▼【修改處 1】修改提問文字，並將變數名稱改為更語意化的 target_successes ▼▼▼
    target_successes = simpledialog.askinteger("輸入", "要湊到幾個「成功」的單字？", minvalue=1)
    if not target_successes:
        print("未輸入處理數量，程式結束。")
        return
    # ▲▲▲【修改結束】▲▲▲

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        print(f"無法開啟 Excel 檔案: {e}")
        return

    headers = [cell.value for cell in ws[1]]
    
    # --- 檢查與設定欄位 ---
    try:
        word_col_idx = headers.index("Words")
        pron_col_idx = headers.index("pronunciation-1")
    except ValueError:
        print("錯誤：Excel 必須包含 'Words' 和 'pronunciation-1' 欄位。")
        return

    # 檢查並新增「複習」欄位
    if "複習" in headers:
        review_col_idx = headers.index("複習")
    else:
        review_col_idx = len(headers)
        ws.cell(row=1, column=review_col_idx + 1, value="複習")
        print("已自動新增 '複習' 欄位。")

    folder = os.path.dirname(excel_path)
    processed_words = {"success": [], "failed": []}
    
    # 找出需要處理的單字
    words_to_process = []
    for row in ws.iter_rows(min_row=2):
        word_cell = row[word_col_idx]
        pron_cell = row[pron_col_idx]
        review_cell = row[review_col_idx] 
        word = word_cell.value
        
        if word and \
           (pron_cell.value is None or str(pron_cell.value).strip() == "") and \
           (review_cell.value != "失敗"):
            words_to_process.append(row)
    
    # ▼▼▼【修改處 2】移除舊的數量限制程式碼 ▼▼▼
    # words_to_process = words_to_process[:count] # <- 這行已被刪除
    # ▲▲▲【修改結束】▲▲▲

    print(f"\n找到 {len(words_to_process)} 個符合條件的單字，開始處理直到成功 {target_successes} 個...")

    # --- 開始處理 ---
    # ▼▼▼【修改處 3】新增成功計數器 ▼▼▼
    success_count = 0
    # ▲▲▲【修改結束】▲▲▲

    for row in words_to_process:
        # ▼▼▼【修改處 4】在每次迴圈開始時，檢查是否已達到目標 ▼▼▼
        if success_count >= target_successes:
            print(f"\n已成功處理 {success_count} 個單字，達到目標數量，處理結束。")
            break # 跳出迴圈
        # ▲▲▲【修改結束】▲▲▲

        word_cell = row[word_col_idx]
        pron_cell = row[pron_col_idx]
        review_cell = row[review_col_idx]
        
        word = str(word_cell.value).strip()
        print(f"處理中 (目前成功 {success_count}/{target_successes} 個): {word}")
        
        fetch_result = fetch_yahoo(word, save_folder=folder)
        
        kk = fetch_result["kk"]
        status = fetch_result["status"]
        error = fetch_result["error"]
        
        review_cell.value = status

        log_entry = {"word": word, "kk": kk, "status": status}

        if status == "成功":
            print(f"  -> 成功 (音標: {kk}, 音檔已下載)")
            pron_cell.value = kk
            pron_cell.font = Font(size=8)
            processed_words["success"].append(log_entry)
            # ▼▼▼【修改處 5】如果成功，計數器加 1 ▼▼▼
            success_count += 1
            # ▲▲▲【修改結束】▲▲▲
        else:
            print(f"  -> 失敗: {error}")
            log_entry["reason"] = error
            processed_words["failed"].append(log_entry)

    # --- 儲存檔案 ---
    try:
        wb.save(excel_path)
        print(f"\n完成！Excel 檔案已儲存至:\n{excel_path}")

        json_path = os.path.join(folder, "processed_words_log.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(processed_words, f, ensure_ascii=False, indent=2)
        print(f"詳細 JSON 記錄檔已儲存至:\n{json_path}")
    except PermissionError:
        print(f"\n錯誤：儲存失敗！請先關閉 Excel 檔案 '{os.path.basename(excel_path)}' 後再試一次。")
    except Exception as e:
        print(f"\n儲存檔案時發生未知錯誤: {e}")


if __name__ == "__main__":
    main()